import locale
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import login as auth_login
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template, render_to_string
from django.views.decorators.http import require_POST
from weasyprint import HTML
from xhtml2pdf import pisa
from django.db.models import Q
from itertools import chain
from django.middleware import csrf
from django.db.models.functions import TruncDay
from django.db.models import Sum, Avg
from .forms import *
from django.http import HttpResponseForbidden
from django.contrib import messages
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.models import Group, User
from django.forms import formset_factory, modelformset_factory
from django.db import transaction
from django.http import JsonResponse
from django.db.models import F, Sum, ExpressionWrapper, IntegerField
from django.db.models import F, Sum, ExpressionWrapper, IntegerField, Value, CharField, Case, When, Q
from django.http import JsonResponse
import requests
import openpyxl
from openpyxl.styles import Alignment, Font
from io import BytesIO
from django.http import HttpResponse
from datetime import datetime
from django.utils.timezone import make_aware
import matplotlib.pyplot as plt
from io import BytesIO
import base64
from django.db.models import Sum, Avg, Count
from django.db.models.functions import TruncDay, TruncMonth

@login_required
def conectado(request):
    request.session['last_activity'] = datetime.timestamp(datetime.now())
    return JsonResponse({'status': 'active'})

def jefe_required(view_func):
    @login_required
    @user_passes_test(lambda u: u.groups.filter(name='staff_jefe').exists())
    def wrapped_view(request, *args, **kwargs):
        return view_func(request, *args, **kwargs)
    return wrapped_view

def bodeguero_required(view_func):
    @login_required
    @user_passes_test(lambda u: u.groups.filter(name='staff_bodega').exists())
    def wrapped_view(request, *args, **kwargs):
        return view_func(request, *args, **kwargs)
    return wrapped_view

def cajero_required(view_func):
    @login_required
    @user_passes_test(lambda u: u.groups.filter(name='staff_vendedor').exists())
    def wrapped_view(request, *args, **kwargs):
        return view_func(request, *args, **kwargs)
    return wrapped_view

def negocio_mayorista_required(view_func):
    @login_required
    def wrapped_view(request, *args, **kwargs):
        
        staff_profile = get_object_or_404(StaffProfile, user=request.user)
        
        if not staff_profile.negocio.is_mayorista:
            return HttpResponseForbidden("Acceso denegado: Este recurso es solo para negocios mayoristas.")
        return view_func(request, *args, **kwargs)
    return wrapped_view

def validar_correo_unico(correo, modelo, negocio=None):
    filtro = {'correo': correo}
    if negocio:
        filtro['negocio'] = negocio

    if modelo.objects.filter(**filtro).exists():
        raise ValidationError(f"El correo '{correo}' ya está registrado en este módulo.")


def obtener_hora_actual():
    url = "https://www.timeapi.io/api/Time/current/zone?timeZone=America/Santiago"
    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()

        # Convertir la fecha y hora de la API en un objeto datetime
        fecha_hora_actual = datetime(
            year=data['year'],
            month=data['month'],
            day=data['day'],
            hour=data['hour'],
            minute=data['minute'],
            second=data['seconds']
        )
        
        # Asegurar que el objeto datetime sea timezone-aware
        fecha_hora_actual = make_aware(fecha_hora_actual)
        return fecha_hora_actual
    except requests.exceptions.RequestException as e:
        # En caso de error, devuelve la hora del servidor
        return make_aware(datetime.now())

def cargar_provincias(request):
    region_id = request.GET.get('region_id')
    provincias = Provincia.objects.filter(region_id=region_id).order_by('nombre')
    return JsonResponse(list(provincias.values('id', 'nombre')), safe=False)

def cargar_comunas(request):
    provincia_id = request.GET.get('provincia_id')
    comunas = Comuna.objects.filter(provincia_id=provincia_id).order_by('nombre')
    return JsonResponse(list(comunas.values('id', 'nombre')), safe=False)

def cargar_ciudades(request):
    comuna_id = request.GET.get('comuna_id')
    ciudades = Ciudad.objects.filter(comuna_id=comuna_id).order_by('nombre')
    return JsonResponse(list(ciudades.values('id', 'nombre')), safe=False)

def user_permissions(request):
    if not request.user.is_authenticated:
        return {}
    return {
        'is_superuser': request.user.is_superuser,
        'es_jefe': request.user.groups.filter(name='staff_jefe').exists(),
        'es_cajero': request.user.groups.filter(name='staff_vendedor').exists(),
        'es_bodeguero': request.user.groups.filter(name='staff_bodega').exists(),
    }

@login_required
@staff_member_required
def detalle_compra(request, compra_id):
    compra = get_object_or_404(Compra, id=compra_id)
    detalles = compra.detalles.all()

    # Seleccionar template basado en `tipo_documento`
    if compra.tipo_documento == "boleta":
        template_name = 'comprobante/compra_exitosa_boleta.html'
    elif compra.tipo_documento == "factura":
        template_name = 'comprobante/compra_exitosa_factura.html'
    else:
        return HttpResponse("Tipo de documento desconocido", status=400)

    return render(request, template_name, {
        'compra': compra,
        'detalles': detalles,
    })

@login_required
@staff_member_required
def home(request):
    user = request.user
    es_jefe = user.groups.filter(name="staff_jefe").exists()
    es_cajero = user.groups.filter(name="staff_vendedor").exists()
    es_bodeguero = user.groups.filter(name="staff_bodega").exists()

    if 'session_warning' in request.session:
        session_warning = request.session['session_warning']
        if session_warning.startswith("logout_warning|"):
            _, message = session_warning.split("|", 1)
            messages.warning(request, message, extra_tags='logout_warning')
        del request.session['session_warning']

    if request.user.is_superuser:
        return redirect('list_admin')

    staff_profile = StaffProfile.objects.get(user=request.user)

    compras = Compra.objects.filter(
        usuario__staffprofile__negocio=staff_profile.negocio
    ).prefetch_related('detalles').order_by('-fecha')

    # Aplicar paginación (10 compras por página)
    paginator = Paginator(compras, 10)
    page_number = request.GET.get('page')  # Obtener el número de página
    compras_page = paginator.get_page(page_number)  # Página actual

    # Cálculo de métricas si el usuario es jefe
    if es_jefe:
        # Total de ventas
        total_ventas = compras.aggregate(Sum('total'))['total__sum'] or 0

        # Ventas del día actual
        today = now().date()
        venta_diaria = compras.filter(fecha__date=today).aggregate(
            total_hoy=Sum('total')
        )['total_hoy'] or 0

        # Calcular el total de costos de entradas de bodega asociadas al negocio
        total_entrada_bodega = EntradaBodegaProducto.objects.filter(
            entrada_bodega__proveedor__negocio=staff_profile.negocio
        ).aggregate(Sum('precio_total'))['precio_total__sum'] or 0

        # Margen de ventas
        margen_ventas = total_ventas - total_entrada_bodega

        # Porcentaje del margen de ventas
        porcentaje_margen = (margen_ventas / total_ventas * 100) if total_ventas > 0 else 0

        # Producto más vendido basado en cantidad
        producto_mas_vendido = DetalleCompra.objects.filter(
            compra__usuario__staffprofile__negocio=staff_profile.negocio
        ).values('producto__nombre').annotate(
            cantidad_total=Sum('cantidad')
        ).order_by('-cantidad_total').first()

        # Producto menos vendido basado en cantidad
        producto_menos_vendido = DetalleCompra.objects.filter(
            compra__usuario__staffprofile__negocio=staff_profile.negocio
        ).values('producto__nombre').annotate(
            cantidad_total=Sum('cantidad')
        ).order_by('cantidad_total').first()

    else:
        total_ventas = venta_diaria = margen_ventas = porcentaje_margen = producto_mas_vendido = producto_menos_vendido = None

    return render(request, 'app/home.html', {
        'staff': user,
        'es_jefe': es_jefe,
        'es_cajero': es_cajero,
        'es_bodeguero': es_bodeguero,
        'compras': compras_page, 
        'total_ventas': total_ventas,
        'venta_diaria': venta_diaria,  # Total ventas día actual
        'margen_ventas': margen_ventas,  # Margen de ventas
        'porcentaje_margen': porcentaje_margen,  # Porcentaje del margen
        'producto_mas_vendido': producto_mas_vendido,
        'producto_menos_vendido': producto_menos_vendido,
    })


def es_ajax(request):
    return request.headers.get('x-requested-with') == 'XMLHttpRequest'

#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
# BOLETA
@login_required
def boleta(request):
    palabra_clave = request.GET.get('buscar', '')
    categoria_filtro = request.GET.get('categoria', None)
    negocio_filtro = request.GET.get('negocio', None)

    if request.user.is_superuser:
        productos = Producto.objects.exclude(precio=0)
        categorias = Categoria.objects.all()
        marcas = Marca.objects.all()
        proveedores = Proveedor.objects.all()
        negocios = Negocio.objects.all()
    else:
        staff_profile = StaffProfile.objects.get(user=request.user)
        productos = Producto.objects.filter(
            almacen__negocio=staff_profile.negocio
        ).filter(Q(precio__gt=0) | Q(is_variable=True))
        categorias = Categoria.objects.filter(negocio=staff_profile.negocio)
        marcas = Marca.objects.filter(negocio=staff_profile.negocio)
        proveedores = Proveedor.objects.filter(negocio=staff_profile.negocio)
        negocios = None

    if palabra_clave:
        productos = productos.filter(Q(nombre__icontains=palabra_clave))

    if categoria_filtro and categoria_filtro.isdigit():
        productos = productos.filter(categoria_id=int(categoria_filtro))

    if request.user.is_superuser and negocio_filtro and negocio_filtro.isdigit():
        productos = productos.filter(almacen__negocio_id=int(negocio_filtro))

    carrito, _ = Carrito.objects.get_or_create(usuario=request.user, tipo="boleta")

    # Aquí se asegura que el cálculo del total utilice los precios actualizados
    carrito_subtotal = sum(item.precio_unitario * item.cantidad for item in carrito.carritoproducto_set.all())
    carrito_descuento_total = sum((item.precio_unitario * item.cantidad * (item.producto.descuento / 100))
                                  for item in carrito.carritoproducto_set.all())
    carrito_iva = (carrito_subtotal - carrito_descuento_total) * 0.19
    carrito_total = carrito_subtotal - carrito_descuento_total 

    return render(request, 'caja/boleta.html', {
        'productos': productos,
        'carrito_items': carrito.carritoproducto_set.all(),
        'carrito_subtotal': carrito_subtotal,
        'carrito_descuento_total': carrito_descuento_total,
        'carrito_iva': carrito_iva,
        'carrito_total': carrito_total,
        'categorias': categorias,
        'marcas': marcas,
        'proveedores': proveedores,
        'negocios': negocios,
        'categoria_filtro': categoria_filtro,
        'negocio_filtro': negocio_filtro,
        'palabra_clave': palabra_clave,
    })

@login_required
def agregar_al_carrito_boleta(request, producto_id):
    producto = get_object_or_404(Producto, producto_id=producto_id)
    carrito, _ = Carrito.objects.get_or_create(usuario=request.user, tipo="boleta")

    carrito_producto, created = CarritoProducto.objects.get_or_create(
        carrito=carrito,
        producto=producto,
        defaults={'cantidad': 1, 'precio_unitario': producto.precio}
    )

    if not created:
        carrito_producto.cantidad += 1
    carrito_producto.save()
    carrito.actualizar_total()
    return redirect('boleta')


@login_required
def agregar_prod_var(request, producto_id):
    producto = get_object_or_404(Producto, producto_id=producto_id)
    carrito, _ = Carrito.objects.get_or_create(usuario=request.user, tipo="boleta")

    if request.method == "POST":
        variable_form = ProductoVariableForm(request.POST)
        precio_form = ProductoPrecioVarForm(request.POST)

        if variable_form.is_valid() and precio_form.is_valid():
            precio_modificado = precio_form.cleaned_data['precio']

            variable_instance = variable_form.save(commit=False)
            variable_instance.producto = producto
            variable_instance.precio_venta = precio_modificado 
            variable_instance.vendido = False
            variable_instance.save()

            carrito_producto, created = CarritoProducto.objects.get_or_create(
                carrito=carrito,
                producto=producto,
                defaults={'precio_unitario': precio_modificado}
            )
            if not created:
                carrito_producto.cantidad +=1
                carrito_producto.precio_unitario = precio_modificado
                carrito_producto.save()

            carrito.actualizar_total()
            return JsonResponse({'success': True})
        else:
            errores = {
                'precio_form_errors': precio_form.errors.as_json(),
                'variable_form_errors': variable_form.errors.as_json()
            }
            return JsonResponse({'success': False, 'errors': errores})
    else:
        html_form = render_to_string(
            'caja/boleta_prod_var.html',
            {'producto': producto, 'variable_form': ProductoVariableForm(), 'precio_form': ProductoPrecioVarForm()},
            request=request,
        )
        return JsonResponse({'html_form': html_form})


@login_required
def boleta_prod_ropa(request, producto_id):
    producto = get_object_or_404(Producto, producto_id=producto_id)
    carrito, _ = Carrito.objects.get_or_create(usuario=request.user, tipo="boleta")

    if request.method == "GET":
        # Obtener las tallas asociadas al producto
        tallas = producto.ropa.all()
        ropa_form = ProductoRopaForm()

        html_form = render_to_string(
            'caja/boleta_prod_ropa.html',
            {"producto": producto, "tallas": tallas, "ropa_form": ropa_form},
            request=request
        )
        return JsonResponse({'html_form': html_form})

    if request.method == "POST":
        ropa_form = ProductoRopaForm(request.POST)
        if ropa_form.is_valid():
            nombre_talla = ropa_form.cleaned_data['nombre_talla']
            unidades = ropa_form.cleaned_data['unidades']

            # Validar stock para la talla seleccionada
            producto_ropa = ProductoRopa.objects.filter(producto=producto, nombre_talla=nombre_talla).first()

            if not producto_ropa:
                return JsonResponse({
                    "success": False,
                    "error": f"La talla {nombre_talla} no existe para el producto seleccionado.",
                })

            if unidades > producto_ropa.unidades:
                return JsonResponse({
                    "success": False,
                    "error": f"No hay suficiente stock para la talla {nombre_talla}.",
                })

            # Agregar las unidades seleccionadas al carrito
            carrito_producto, created = CarritoProducto.objects.get_or_create(
                carrito=carrito,
                producto=producto,
                defaults={'cantidad': 0, 'precio_unitario': producto.precio}
            )
            carrito_producto.cantidad += unidades
            carrito_producto.save()

            # Actualizar las unidades en ProductoRopa
            producto_ropa.unidades -= unidades
            producto_ropa.save()

            # Actualizar el total del carrito
            carrito.actualizar_total()

            return JsonResponse({
                "success": True,
                "message": f"{unidades} unidades del producto '{producto.nombre}' (talla {nombre_talla}) añadidas al carrito."
            })

        # Manejo de errores en el formulario
        return JsonResponse({"success": False, "errors": ropa_form.errors})

@require_POST
@login_required
def eliminar_del_carrito_boleta(request, item_id):
    carrito_producto = get_object_or_404(CarritoProducto, id=item_id, carrito__usuario=request.user, carrito__tipo="boleta")

    carrito_producto.delete()
    carrito_producto.carrito.actualizar_total()
    
    return redirect('boleta')


@login_required
def confirmar_compra_boleta(request):
    if request.method == 'POST':
        # Identificar si la solicitud es para cancelar la venta
        if 'cancelar' in request.POST:
            carrito = Carrito.objects.filter(usuario=request.user, tipo="boleta").first()
            if carrito:
                carrito.carritoproducto_set.all().delete()
                carrito.actualizar_total()

            messages.success(request, "El carrito ha sido vaciado.")
            return redirect('boleta')

        correo = request.POST.get('correo', '').strip()
        medio_pago = request.POST.get('medio_pago', '')

        carrito = get_object_or_404(Carrito, usuario=request.user, tipo="boleta")
        carrito_items = carrito.carritoproducto_set.all()

        if carrito_items:
            staff_profile = StaffProfile.objects.get(user=request.user)
            negocio = staff_profile.negocio

            perfil_cliente = None
            if correo:
                perfil_cliente, _ = PerfilClientes.objects.get_or_create(correo=correo)

            # Calcular los totales considerando precio temporal
            subtotal = sum(item.precio_unitario * item.cantidad for item in carrito_items) 
            descuento_total = sum(
                (item.precio_unitario * item.cantidad * item.producto.descuento / 100) for item in carrito_items
            )
            iva_total = (subtotal - descuento_total) * 0.19
            # subtotal = subtotal - iva_total
            # corrigiendo logica de calculo de valores
            total = subtotal - descuento_total 

            # Registrar la compra
            compra = Compra.objects.create(
                usuario=request.user,
                negocio=negocio,
                subtotal=subtotal,
                descuento_total=descuento_total,
                iva_total=iva_total,
                total=total,
                nombre_staff=request.user.get_full_name(),
                correo=perfil_cliente.correo if perfil_cliente else None,
                medio_pago=medio_pago,
                tipo_documento='boleta',
                fecha=obtener_hora_actual()
            )

        # Guardar cada producto, incluyendo productos variables
        for item in carrito_items:
            producto = item.producto

            
            if producto.ropa.exists():
                    # Buscar la información de las tallas asociadas desde el formulario de ropa
                    tallas_ropa = producto.ropa.all()
                    cantidad_restante = item.cantidad

                    for talla in tallas_ropa:
                        if cantidad_restante <= 0:
                            break

                        if talla.unidades > 0:
                            unidades_a_descontar = min(cantidad_restante, talla.unidades)
                            talla.unidades -= unidades_a_descontar
                            talla.save()
                            cantidad_restante -= unidades_a_descontar

                    if cantidad_restante > 0:
                        messages.error(request, f"No hay suficientes unidades disponibles para el producto {producto.nombre}.")
                        return redirect('boleta')


            if producto.is_variable:
                # Obtener el registro de ProductoVariable para mantener la referencia
                    producto_variable = ProductoVariable.objects.filter(producto=producto).first()
                    if producto_variable:
                        
                        producto_variable.cantidad = (producto_variable.cantidad or 0) + item.cantidad  # Actualizar la cantidad vendida
                        producto_variable.vendido = True  
                        producto_variable.save()

                        DetalleCompra.objects.create(
                            compra=compra,
                            producto=producto,
                            cantidad=item.cantidad,
                            precio_unitario=item.precio_unitario
                        )
            else:
                if producto.stock >= item.cantidad:
                    producto.stock -= item.cantidad
                    producto.save()
                    DetalleCompra.objects.create(
                        compra=compra,
                        producto=producto,
                        cantidad=item.cantidad,
                        precio_unitario=item.precio_unitario
                    )
                else:
                    messages.error(request, f"No hay suficiente stock para {producto.nombre}.")
                    return redirect('boleta')

            # Limpiar el carrito
            carrito.carritoproducto_set.all().delete()
            carrito.actualizar_total()

            # Generar PDF y redirigir
            if correo:
                template_path = 'comprobante/boleta_pdf.html'
                context = {'compra': compra, 'detalles': compra.detalles.all()}
                pdf = generar_pdf(template_path, context)
                if pdf:
                    enviar_correo(correo, '¡Su compra ha sido exitosa!', 'Gracias por tu compra.', pdf, 'boleta.pdf')

            # Redirigir a la boleta generada
            return redirect('compra_exitosa_boleta', compra_id=compra.id)
        else:
            messages.warning(request, 'El carrito está vacío.')
            return redirect('boleta')
    else:
        return redirect('boleta')



def generar_pdf(template_path, context):
    template = get_template(template_path)
    html = template.render(context)
    result = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=result)
    if not pisa_status.err:
        return result.getvalue()
    return None

def enviar_correo(destinatario, asunto, mensaje, archivo, nombre_archivo):
    email = EmailMessage(asunto, mensaje, 'contacto@atlasgestion.cl', [destinatario])
    email.attach(nombre_archivo, archivo, 'application/pdf')
    email.send()

@require_POST
@login_required
def restar_producto_boleta(request, producto_id):
    producto = get_object_or_404(Producto, producto_id=producto_id)
    carrito = get_object_or_404(Carrito, usuario=request.user, tipo="boleta")
    carrito_producto = get_object_or_404(CarritoProducto, carrito=carrito, producto=producto)

    if carrito_producto.cantidad > 1:
        carrito_producto.cantidad -= 1
        carrito_producto.save()
    else:
        carrito_producto.delete()

    carrito.actualizar_total()
    return redirect('boleta')

@login_required
@permission_required('app.add_producto', raise_exception=True)
def reg_prod_boleta(request):
    staff_profile = get_object_or_404(StaffProfile, user=request.user)
    negocio = staff_profile.negocio

    if request.method == "GET":
        form = ProductoManualForm()
        return render(request, 'caja/reg_prod_boleta.html', {
            'form': form,
        })

    elif request.method == "POST":
        form = ProductoManualForm(request.POST)

        if form.is_valid():
            producto = form.save(commit=False)
            producto.descuento = producto.descuento or 0
            producto.precio = producto.precio or 0
            producto.precio_mayorista = producto.precio_mayorista or 0
            producto.descuento_mayorista = producto.descuento_mayorista or 0
            producto.tasa_ila = 0  # ILA no aplica
            producto.estado = "ingresado_manual"
            producto.almacen = Almacen.objects.filter(negocio=staff_profile.negocio).first()

            if not producto.almacen:
                messages.error(request, "No se encontró un almacén asociado al negocio.")
                return JsonResponse({"success": False}, status=400)

            producto.save()
            # Agregar al carrito
            carrito, _ = Carrito.objects.get_or_create(usuario=request.user, tipo="boleta")
            carrito_producto, created = CarritoProducto.objects.get_or_create(
                carrito=carrito,
                producto=producto,
                defaults={
                    'cantidad': 1,
                    'precio_unitario': producto.precio,
                },
            )
            if not created:
                carrito_producto.cantidad += 1
                carrito_producto.save()

            carrito.actualizar_total()

            messages.success(request, f'El producto "{producto.nombre}" ha sido registrado y agregado al carrito de boleta.')
            return JsonResponse({"success": True})
        else:
            # Registrar errores en el sistema de mensajes
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")

            return JsonResponse({"success": False}, status=400)


@login_required
@permission_required('app.change_producto', raise_exception=True)
def actualizar_descuento_boleta(request, producto_id):
    producto = get_object_or_404(Producto, producto_id=producto_id)
    
    if request.method == 'POST':
        nuevo_descuento = request.POST.get('descuento', 0)  
        
        if nuevo_descuento:
            producto.descuento = int(nuevo_descuento)
        
        producto.save()
        messages.success(request, f"El descuento de '{producto.nombre}' han sido actualizados.")

    return redirect('boleta')

@login_required
def vaciar_carrito_boleta(request):
    carrito = Carrito.objects.filter(usuario=request.user, tipo="boleta").first()
    if carrito:
        carrito.carritoproducto_set.all().delete()  # Elimina todos los productos del carrito
        carrito.actualizar_total()  # Resetea el total del carrito
    messages.success(request, "El carrito ha sido vaciado.")
    return redirect('boleta')  # Redirige a la página de boleta
#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
# FACTURA
@login_required
@negocio_mayorista_required
def factura(request):
    palabra_clave = request.GET.get('buscar', '')
    categoria_filtro = request.GET.get('categoria', None)
    negocio_filtro = request.GET.get('negocio', None)
    
    if request.user.is_superuser:
        productos = Producto.objects.exclude(precio_mayorista=0)
        categorias = Categoria.objects.all()
        marcas = Marca.objects.all()
        proveedores = Proveedor.objects.all()
        negocios = Negocio.objects.all()
    else:
        staff_profile = StaffProfile.objects.get(user=request.user)
        productos = Producto.objects.filter(
        almacen__negocio=staff_profile.negocio
        ).filter(Q(precio__gt=0) | Q(is_variable=True))
        categorias = Categoria.objects.filter(negocio=staff_profile.negocio)
        marcas = Marca.objects.filter(negocio=staff_profile.negocio)
        proveedores = Proveedor.objects.filter(negocio=staff_profile.negocio)

        empresas = PerfilClienteEmpresa.objects.filter(negocio=staff_profile.negocio, activo=True) # Filtramos las empresas activas
        empresa_form = PerfilClienteEmpresaForm()
        negocios = None

    if palabra_clave:
        productos = productos.filter(Q(nombre__icontains=palabra_clave))

    if categoria_filtro and categoria_filtro.isdigit():
        productos = productos.filter(categoria_id=int(categoria_filtro))

    if request.user.is_superuser and negocio_filtro and negocio_filtro.isdigit():
        productos = productos.filter(almacen__negocio_id=int(negocio_filtro))

    carrito, _ = Carrito.objects.get_or_create(usuario=request.user, tipo="factura")
    carrito.actualizar_total_mayorista()  

    regiones = Region.objects.all()
    
    return render(request, 'caja/factura.html', {
        'productos': productos,
        'carrito_items': carrito.carritoproducto_set.all(),
        'carrito_subtotal': carrito.subtotal,
        'carrito_descuento_total': carrito.descuento_total,
        'carrito_iva': carrito.iva_total,
        'carrito_total': carrito.total,
        'categorias': categorias,
        'marcas': marcas,
        'proveedores': proveedores,
        'negocios': negocios,
        'categoria_filtro': categoria_filtro,
        'negocio_filtro': negocio_filtro,
        'palabra_clave': palabra_clave,
        'empresas': empresas,
        'empresa_form': empresa_form,
        'regions': regiones
    })



@login_required
def agregar_al_carrito_factura(request, producto_id):
    producto = get_object_or_404(Producto, producto_id=producto_id)
    carrito, _ = Carrito.objects.get_or_create(usuario=request.user, tipo="factura")

    carrito_producto, created = CarritoProducto.objects.get_or_create(
        carrito=carrito,
        producto=producto,
        defaults={'cantidad': 1, 'precio_unitario': producto.precio_mayorista}
    )

    if not created:
        carrito_producto.cantidad += 1
    carrito_producto.save()
    carrito.actualizar_total_mayorista()
    return redirect('factura')


@require_POST
@login_required
def eliminar_del_carrito_factura(request, item_id):
    carrito_producto = get_object_or_404(CarritoProducto, id=item_id, carrito__usuario=request.user, carrito__tipo="factura")
    carrito_producto.delete()
    carrito_producto.carrito.actualizar_total_mayorista()
    return redirect('factura')

from django.core.mail import EmailMessage
from xhtml2pdf import pisa
from io import BytesIO

def generar_pdf(template_path, context):
    template = get_template(template_path)
    html = template.render(context)
    result = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=result)
    if not pisa_status.err:
        return result.getvalue()
    return None

def enviar_correo(destinatario, asunto, mensaje, archivo, nombre_archivo):
    email = EmailMessage(asunto, mensaje, 'contacto@atlasgestion.cl', [destinatario])
    email.attach(nombre_archivo, archivo, 'application/pdf')
    email.send()

def enviar_correo_datos(destinatario, asunto, mensaje, archivo=None, nombre_archivo=''):
    email = EmailMessage(asunto, mensaje, 'contacto@atlasgestion.cl', [destinatario])
    if archivo and nombre_archivo: 
        email.attach(nombre_archivo, archivo, 'application/pdf')
    email.send()


@login_required
def buscar_correo(request):
    term = request.GET.get('term', '').strip()
    correos = PerfilClienteEmpresa.objects.filter(correo__icontains=term).values('id', 'correo')[:10]

    # Crear correo automáticamente si no existe y es válido
    if not correos.exists() and '@' in term:  # Valida un formato básico de correo
        nuevo_perfil, created = PerfilClienteEmpresa.objects.get_or_create(correo=term)
        correos = [{'id': nuevo_perfil.id, 'correo': nuevo_perfil.correo}]

    results = [{'id': correo['id'], 'label': correo['correo'], 'value': correo['correo']} for correo in correos]
    return JsonResponse(results, safe=False)


@login_required
def confirmar_compra_factura(request):
    staff_profile = StaffProfile.objects.get(user=request.user)
    negocio = staff_profile.negocio
    if not negocio:
        messages.error(request, "No se puede procesar la factura porque el staff no tiene un negocio asociado.")
        return redirect('factura')

    if request.method == 'POST':
        medio_pago = request.POST.get('medio_pago', '')
        glosa = request.POST.get('glosa', '').strip()
        correo = request.POST.get('correo', '').strip()

        carrito = get_object_or_404(Carrito, usuario=request.user, tipo="factura")
        carrito_items = carrito.carritoproducto_set.all()

        if carrito_items:
            perfil_cliente = None
            if correo:
                correo = correo.lower().strip()  # Normalizar el correo
                try:
                    perfil_cliente = PerfilClienteEmpresa.objects.get(correo=correo, negocio=negocio)
                except PerfilClienteEmpresa.DoesNotExist:
                    
                    perfil_cliente = PerfilClienteEmpresa(
                        correo=correo,
                        negocio=negocio
                    )
                    perfil_cliente.full_clean()  # Validación antes de guardar
                    perfil_cliente.save()

            # Obtener la fecha y hora actual
            fecha_hora_actual = obtener_hora_actual()

            # Calcular subtotal, descuentos y IVA
            subtotal = sum((item.producto.precio_mayorista / 1.19) * item.cantidad for item in carrito_items)
            descuento_total = sum(
                item.producto.precio_mayorista * item.cantidad * item.producto.descuento_mayorista / 100 for item in carrito_items
            )
            iva_total = (subtotal - descuento_total) * 0.19
            total = subtotal - descuento_total + iva_total

            negocio = staff_profile.negocio 
            compra = Compra.objects.create(
                usuario=request.user,
                negocio=negocio, 
                subtotal=subtotal,
                descuento_total=descuento_total,
                iva_total=iva_total,
                total=total,
                nombre_staff=request.user.get_full_name(),
                correo=perfil_cliente.correo if perfil_cliente else None,
                medio_pago=medio_pago,
                glosa=glosa,
                tipo_documento='factura', 
                fecha=fecha_hora_actual
            )

            # Procesar cada item en el carrito
            for item in carrito_items:
                producto = item.producto
                if producto.stock >= item.cantidad:
                    producto.stock -= item.cantidad
                    producto.save()

                    DetalleCompra.objects.create(
                        compra=compra,
                        producto=producto,
                        cantidad=item.cantidad,
                        precio_unitario=producto.precio_mayorista
                    )
                else:
                    messages.error(request, f"No hay suficiente stock para {producto.nombre}.")
                    return redirect('factura')

            # Limpiar el carrito después de la compra
            carrito.carritoproducto_set.all().delete()
            carrito.actualizar_total_mayorista()

            # Enviar correo con factura en PDF
            if correo:
                template_path = 'comprobante/factura_pdf.html'
                context = {'compra': compra, 'detalles': compra.detalles.all()}
                pdf = generar_pdf(template_path, context)
                if pdf:
                    enviar_correo(
                        correo, 
                        '¡Su compra ha sido exitosa!',
                        'Hola, a continuación te adjuntamos la Factura electrónica asociada a tu compra, para que esté disponible dónde y cuándo quieras. Guarda esta factura e imprímela sólo de ser necesario. ¡Cuidemos juntos nuestro planeta!.',
                        pdf, 
                        'factura.pdf'
                    )

            messages.success(request, 'La compra ha sido confirmada.')
            return redirect('compra_exitosa_factura', compra_id=compra.id)
        else:
            messages.warning(request, 'El carrito está vacío.')
            return redirect('factura')


@require_POST
@login_required
def restar_producto_factura(request, producto_id):
    producto = get_object_or_404(Producto, producto_id=producto_id)
    carrito = get_object_or_404(Carrito, usuario=request.user, tipo="factura")
    carrito_producto = get_object_or_404(CarritoProducto, carrito=carrito, producto=producto)

    if carrito_producto.cantidad > 1:
        carrito_producto.cantidad -= 1
        carrito_producto.save()
    else:
        carrito_producto.delete()

    carrito.actualizar_total_mayorista()
    return redirect('factura')

@login_required
@permission_required('app.add_producto', raise_exception=True)
def reg_prod_factura(request):
    staff_profile = get_object_or_404(StaffProfile, user=request.user)

    if request.method == "GET":
        form = ProductoManualFacturaForm()
        return render(request, 'caja/reg_prod_factura.html', {'form': form})

    elif request.method == "POST":
        nombre_producto = request.POST.get('nombre', '').strip()
        producto_id = request.POST.get('producto_id', None)

        # Verificar si el producto ya existe
        producto = None
        if producto_id:
            producto = Producto.objects.filter(id=producto_id).first()
        elif nombre_producto:
            producto = Producto.objects.filter(nombre__iexact=nombre_producto).first()

        if producto:
            # Si el producto existe, agregarlo al carrito
            carrito, _ = Carrito.objects.get_or_create(usuario=request.user, tipo="factura")
            carrito_producto, created = CarritoProducto.objects.get_or_create(
                carrito=carrito,
                producto=producto,
                defaults={'cantidad': 1, 'precio_unitario': producto.precio_mayorista},
            )
            if not created:
                carrito_producto.cantidad += 1
                carrito_producto.save()
            carrito.actualizar_total_mayorista()

            return JsonResponse({"success": True, "message": f"Producto '{producto.nombre}' agregado al carrito."})
        else:
            # Si el producto no existe, crearlo
            form = ProductoManualFacturaForm(request.POST)
            if form.is_valid():
                nuevo_producto = form.save(commit=False)
                nuevo_producto.almacen = Almacen.objects.filter(negocio=staff_profile.negocio).first()
                nuevo_producto.estado = "ingresado_manual"
                nuevo_producto.save()

                # Agregar al carrito
                carrito, _ = Carrito.objects.get_or_create(usuario=request.user, tipo="factura")
                carrito_producto = CarritoProducto.objects.create(
                    carrito=carrito,
                    producto=nuevo_producto,
                    cantidad=1,
                    precio_unitario=nuevo_producto.precio_mayorista
                )
                carrito.actualizar_total_mayorista()

                return JsonResponse({"success": True, "message": f"Producto '{nuevo_producto.nombre}' creado y agregado al carrito."})
            else:
                return JsonResponse({"success": False, "message": "Datos inválidos en el formulario", "errors": form.errors}, status=400)

    return JsonResponse({"success": False, "message": "Método no permitido"}, status=405)

@login_required
def buscar_producto(request):
    term = request.GET.get('term', '').strip()
    productos = Producto.objects.filter(nombre__icontains=term).values('id', 'nombre')[:10]
    results = [{'id': producto['id'], 'label': producto['nombre'], 'value': producto['nombre']} for producto in productos]
    return JsonResponse(results, safe=False)



@login_required
@permission_required('app.change_producto', raise_exception=True)
def actualizar_descuento_factura(request, producto_id):
    producto = get_object_or_404(Producto, producto_id=producto_id)
    
    if request.method == 'POST':
        nuevo_descuento = request.POST.get('descuento', 0)  
        
        if nuevo_descuento:
            producto.descuento_mayorista = int(nuevo_descuento)
        
        producto.save()
        messages.success(request, f"El descuento de '{producto.nombre}' han sido actualizados.")


    return redirect('factura')

@login_required
def vaciar_carrito_factura(request):
    carrito = Carrito.objects.filter(usuario=request.user, tipo="factura").first()
    if carrito:
        carrito.carritoproducto_set.all().delete()  
        carrito.actualizar_total_mayorista() 
    messages.success(request, "El carrito ha sido vaciado.")
    return redirect('factura')  

from django.shortcuts import get_object_or_404, render
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import Producto, ProductoRopa, Carrito, CarritoProducto
from .forms import ProductoRopaForm, ProductoVariableForm

@login_required
def factura_prod_ropa(request, producto_id):
    producto = get_object_or_404(Producto, producto_id=producto_id)
    carrito, _ = Carrito.objects.get_or_create(usuario=request.user, tipo="factura")

    if request.method == "GET":
        tallas = producto.ropa.all()
        ropa_form = ProductoRopaForm()
        html_form = render_to_string(
            'caja/factura_prod_ropa.html',
            {"producto": producto, "tallas": tallas, "ropa_form": ropa_form},
            request=request
        )
        return JsonResponse({'html_form': html_form})

    if request.method == "POST":
        ropa_form = ProductoRopaForm(request.POST)
        if ropa_form.is_valid():
            nombre_talla = ropa_form.cleaned_data['nombre_talla']
            unidades = ropa_form.cleaned_data['unidades']

            producto_ropa = get_object_or_404(ProductoRopa, producto=producto, nombre_talla=nombre_talla)
            if unidades > producto_ropa.unidades:
                return JsonResponse({
                    "success": False,
                    "error": f"No hay suficiente stock para la talla {nombre_talla}.",
                })

            carrito_producto, created = CarritoProducto.objects.get_or_create(
                carrito=carrito,
                producto=producto,
                defaults={'cantidad': 0, 'precio_unitario': producto.precio_mayorista}
            )
            carrito_producto.cantidad += unidades
            carrito_producto.save()

            carrito.actualizar_total()

            return JsonResponse({
                "success": True,
                "message": f"{unidades} unidades del producto '{producto.nombre}' (talla {nombre_talla}) añadidas al carrito."
            })

        return JsonResponse({"success": False, "errors": ropa_form.errors})


@login_required
def factura_prod_var(request, producto_id):
    producto = get_object_or_404(Producto, producto_id=producto_id)
    carrito, _ = Carrito.objects.get_or_create(usuario=request.user, tipo="factura")

    if request.method == "POST":
        variable_form = ProductoVariableForm(request.POST)
        precio_form = ProductoPrecioVarForm(request.POST)
        if variable_form.is_valid():
            precio_modificado = variable_form.cleaned_data['precio']

            variable_instance = variable_form.save(commit=False)
            variable_instance.producto = producto
            variable_instance.precio_venta = precio_modificado 
            variable_instance.vendido = False
            variable_instance.save()

            carrito_producto, created = CarritoProducto.objects.get_or_create(
                carrito=carrito,
                producto=producto,
                defaults={'cantidad': 1, 'precio_unitario': precio_modificado}
            )
            if not created:
                carrito_producto.cantidad +=1
                carrito_producto.precio_unitario = precio_modificado
                carrito_producto.save()

            carrito.actualizar_total_mayorista()
            return JsonResponse({'success': True})
        else:
            errores = {
                'precio_form_errors': precio_form.errors.as_json(),
                'variable_form_errors': variable_form.errors.as_json()
            }
            return JsonResponse({'success': False, 'errors': errores})
    else:
        html_form = render_to_string(
            'caja/factura_prod_var.html',
            {'producto': producto, 'variable_form': ProductoVariableForm(), 'precio_form': ProductoPrecioVarForm()},
            request=request,
        )
        return JsonResponse({'html_form': html_form})
#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
# VENTA EXISTOSA
@login_required
def compra_exitosa_boleta(request, compra_id):
    compra = get_object_or_404(Compra, id=compra_id)

    detalles = compra.detalles.all()

    if 'pdf' in request.GET:
        template_path = 'comprobante/boleta_pdf.html'
        context = {
            'compra': compra,
            'detalles': detalles
        }

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="comprobante_boleta.pdf"'

        template = get_template(template_path)
        html = template.render(context)
        pisa_status = pisa.CreatePDF(html, dest=response)

        if pisa_status.err:
            return HttpResponse('Hubo un error al generar el PDF <pre>%s</pre>' % html)

        return response

    return render(request, 'comprobante/compra_exitosa_boleta.html', {
        'compra': compra,
        'detalles': detalles
    })


@login_required
def compra_exitosa_factura(request, compra_id):
    compra = get_object_or_404(Compra, id=compra_id)

    detalles = compra.detalles.all()

    if 'pdf' in request.GET:
        template_path = 'comprobante/factura_pdf.html'
        context = {
            'compra': compra,
            'detalles': detalles
        }

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="comprobante_factura.pdf"'

        template = get_template(template_path)
        html = template.render(context)
        pisa_status = pisa.CreatePDF(html, dest=response)

        if pisa_status.err:
            return HttpResponse('Hubo un error al generar el PDF <pre>%s</pre>' % html)

        return response

    return render(request, 'comprobante/compra_exitosa_factura.html', {
        'compra': compra,
        'detalles': detalles
    })


def confirmar_compra_invitado(request):
    email = request.POST.get('email_invitado')
    carrito = get_object_or_404(Carrito, usuario=request.user)
    carrito_items = carrito.carritoproducto_set.all()

    if carrito_items:
        compra = Compra.objects.create(usuario=None, total=carrito.total, correo=email)

        for item in carrito_items:
            DetalleCompra.objects.create(
                compra=compra,
                producto=item.producto,
                cantidad=item.cantidad,
                precio_unitario=item.producto.precio
            )

        carrito.carritoproducto_set.all().delete()
        carrito.total = 0
        carrito.save()

        return redirect('compra_exitosa', compra_id=compra.id)

    return redirect('home')


#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
# LOGIN Y LOGOUT

def login(request):
    if 'session_warning' in request.session:
        session_warning = request.session['session_warning']
        if session_warning.startswith("logout_warning|"):
            _, message = session_warning.split("|", 1)
            messages.warning(request, message, extra_tags='logout_warning')
        del request.session['session_warning']

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            auth_login(request, user)
            return redirect('home')
    else:
        form = AuthenticationForm()

    return render(request, 'registration/login.html', {'form': form})
@require_POST
def logoutView(request):
    logout(request) 
    return redirect('login') 
#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
# CRUD PRODUCTO
@login_required
def buscar_categoria(request):
    term = request.GET.get('term', '').strip()

    try:
        staff_profile = StaffProfile.objects.get(user=request.user)
        negocio = staff_profile.negocio
    except StaffProfile.DoesNotExist:
        return JsonResponse([], safe=False)

    categorias = Categoria.objects.filter(
        nombre__icontains=term,
        negocio=negocio 
    ).values('id', 'nombre')[:10]
    
    results = [{'id': categoria['id'], 'label': categoria['nombre'], 'value': categoria['nombre']} for categoria in categorias]
    return JsonResponse(results, safe=False)

@login_required
def buscar_marca(request):
    term = request.GET.get('term', '').strip()

    try:
        staff_profile = StaffProfile.objects.get(user=request.user)
        negocio = staff_profile.negocio
    except StaffProfile.DoesNotExist:
        return JsonResponse([], safe=False)

    marcas = Marca.objects.filter(
        nombre__icontains=term,
        negocio=negocio
    ).values('id', 'nombre')[:10]
    
    results = [{'id': marca['id'], 'label': marca['nombre'], 'value': marca['nombre']} for marca in marcas]
    return JsonResponse(results, safe=False)


@login_required
@permission_required('app.add_producto', raise_exception=True)
def add_prod_modal(request):
    staff_profile = StaffProfile.objects.get(user=request.user)
    negocio = staff_profile.negocio

    FormularioProducto = ProductoFormMayorista if negocio.is_mayorista else ProductoFormMinorista

    if request.method == 'POST':
        producto_form = FormularioProducto(request.POST)

        ropa_forms_data = []

        if 'ropaCheckbox' in request.POST:
            ropa_total_forms = int(request.POST.get('ropa-TOTAL_FORMS', 0))
            for i in range(ropa_total_forms):
                talla = request.POST.get(f'form-{i}-nombre_talla')
                unidades = request.POST.get(f'form-{i}-unidades')
                if talla and unidades:
                    ropa_forms_data.append({'nombre_talla': talla, 'unidades': unidades})

        is_variable = 'variableCheckbox' in request.POST

        if producto_form.is_valid():
            producto = producto_form.save(commit=False)
            producto.is_variable = is_variable

            categoria_id = request.POST.get('categoria_id')
            categoria_nombre = request.POST.get('categoria_nombre', '').strip()
            if categoria_id:
                categoria = Categoria.objects.get(id=categoria_id)
            elif categoria_nombre:
                categoria, _ = Categoria.objects.get_or_create(nombre=categoria_nombre, defaults={'negocio': negocio})
            else:
                categoria = None
            producto.categoria = categoria

            marca_id = request.POST.get('marca_id')
            marca_nombre = request.POST.get('marca_nombre', '').strip()
            if marca_id:
                marca = Marca.objects.get(id=marca_id)
            elif marca_nombre:
                marca, _ = Marca.objects.get_or_create(nombre=marca_nombre, defaults={'negocio': negocio})
            else:
                marca = None
            producto.marca = marca

            producto.descuento = producto.descuento or 0
            producto.precio = producto.precio or 0
            producto.precio_mayorista = producto.precio_mayorista or 0
            producto.descuento_mayorista = producto.descuento_mayorista or 0
            producto.almacen = Almacen.objects.filter(negocio=staff_profile.negocio).first()

            producto.save()

            for ropa_data in ropa_forms_data:
                ProductoRopa.objects.create(
                    producto=producto,
                    nombre_talla=ropa_data['nombre_talla'],
                    unidades=int(ropa_data['unidades'])
                )

            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'errors': producto_form.errors})

    else:
        producto_form = FormularioProducto()
        ropa_form = ProductoRopaForm()

    html_form = render_to_string(
        'producto/add_prod_modal_form.html',
        {
            'producto_form': producto_form,
            'ropa_form': ropa_form
        },
        request=request,
    )
    return JsonResponse({'html_form': html_form})

@login_required
@permission_required('app.change_producto', raise_exception=True)
def mod_prod_modal(request, producto_id):
    producto = get_object_or_404(Producto, producto_id=producto_id)
    staff_profile = StaffProfile.objects.get(user=request.user)
    negocio = staff_profile.negocio

    FormularioProducto = ProductoFormMayorista if negocio.is_mayorista else ProductoFormMinorista

    # Detectar subformularios asociados al producto
    ropa_forms = ProductoRopa.objects.filter(producto=producto)
    unidades_form = ProductoUnidadesForm(instance=producto.unidades) if hasattr(producto, 'unidades') else None

    if request.method == 'POST':
        producto_form = FormularioProducto(request.POST, instance=producto)

        ropa_data = []
        if ropa_forms.exists():
            ropa_total_forms = int(request.POST.get('ropa-TOTAL_FORMS', 0))
            for i in range(ropa_total_forms):
                talla = request.POST.get(f'form-{i}-nombre_talla')
                unidades = request.POST.get(f'form-{i}-unidades')
                if talla and unidades:
                    ropa_data.append({'nombre_talla': talla, 'unidades': unidades})

        if producto_form.is_valid():
            producto = producto_form.save(commit=False)

            # Manejo de categoría
            categoria_id = request.POST.get('categoria_id')
            categoria_nombre = request.POST.get('categoria_nombre', '').strip()
            if categoria_id:
                categoria = Categoria.objects.get(id=categoria_id)
            elif categoria_nombre:
                categoria, _ = Categoria.objects.get_or_create(nombre=categoria_nombre, defaults={'negocio': negocio})
            else:
                categoria = None
            producto.categoria = categoria

            # Manejo de marca
            marca_id = request.POST.get('marca_id')
            marca_nombre = request.POST.get('marca_nombre', '').strip()
            if marca_id:
                marca = Marca.objects.get(id=marca_id)
            elif marca_nombre:
                marca, _ = Marca.objects.get_or_create(nombre=marca_nombre, defaults={'negocio': negocio})
            else:
                marca = None
            producto.marca = marca

            producto.save()

            # Guardar datos de subformularios
            ProductoRopa.objects.filter(producto=producto).delete()
            for ropa in ropa_data:
                ProductoRopa.objects.create(
                    producto=producto,
                    nombre_talla=ropa['nombre_talla'],
                    unidades=int(ropa['unidades'])
                )

            if unidades_form and unidades_form.is_valid():
                unidades = unidades_form.save(commit=False)
                unidades.producto = producto
                unidades.save()

            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'errors': producto_form.errors})

    else:
        producto_form = FormularioProducto(instance=producto)

    html_form = render_to_string('producto/mod_prod_modal_form.html', {
        'producto_form': producto_form,
        'producto': producto,
        'ropa_forms': ropa_forms,
        'unidades_form': unidades_form,
        'categoria_nombre': producto.categoria.nombre if producto.categoria else '',
        'categoria_id': producto.categoria.id if producto.categoria else '',
        'marca_nombre': producto.marca.nombre if producto.marca else '',
        'marca_id': producto.marca.id if producto.marca else '',
    }, request=request)
    return JsonResponse({'html_form': html_form})


from django.db import IntegrityError

@login_required
def list_prod(request):
    palabra_clave = request.GET.get('buscar', '')
    estado_filtro = request.GET.get('estado', None)
    negocio_filtro = request.GET.get('negocio', None)
    categoria_filtro = request.GET.get('categoria', None)

    error_message = None
    marca_form = MarcaForm()
    categoria_form = CategoriaForm()

    if request.user.is_superuser:
        negocios = Negocio.objects.all()
        productos = Producto.objects.exclude(estado='descontinuado').order_by(
            models.Case(
                models.When(estado='disponible', then=0),
                models.When(estado='sin_stock', then=1),
                models.When(estado='registrado_reciente', then=2),
                models.When(estado='ingresado_manual', then=3),
                default=4,
                output_field=models.IntegerField(),
            ),
            'nombre'
        )
        negocio_nombre = None
        if negocio_filtro:
            try:
                negocio_filtro = int(negocio_filtro)
                productos = productos.filter(almacen__negocio_id=negocio_filtro)
                negocio_nombre = Negocio.objects.get(id=negocio_filtro).nombre
            except (ValueError, Negocio.DoesNotExist):
                negocio_filtro = None

        if estado_filtro:
            productos = productos.filter(estado=estado_filtro)

        if categoria_filtro:
            try:
                categoria_filtro = int(categoria_filtro)
                productos = productos.filter(categoria_id=categoria_filtro)
            except (ValueError, Categoria.DoesNotExist):
                categoria_filtro = None

        categorias = Categoria.objects.all()
        marcas = Marca.objects.all()
    else:
        staff_profile = StaffProfile.objects.get(user=request.user)
        negocio = staff_profile.negocio
        productos = Producto.objects.filter(almacen__negocio=negocio).exclude(estado='ingresado_manual').order_by(
            models.Case(
                models.When(estado='disponible', then=0),
                models.When(estado='sin_stock', then=1),
                models.When(estado='registrado_reciente', then=2),
                #models.When(estado='ingresado_manual', then=3),
                default=4,
                output_field=models.IntegerField(),
            ),
            'nombre'
        )
        categorias = Categoria.objects.filter(negocio=negocio)
        marcas = Marca.objects.filter(negocio=negocio)
        negocio_nombre = negocio.nombre
        negocios = None

        if palabra_clave:
            productos = productos.filter(Q(nombre__icontains=palabra_clave))

        if estado_filtro:
            productos = productos.filter(estado=estado_filtro)

        if categoria_filtro and categoria_filtro.isdigit():
            productos = productos.filter(categoria_id=int(categoria_filtro))

        if request.method == 'POST':
            form_type = request.POST.get('form_type')
            
            if form_type == "marca":
                # Manejo del formulario de Marca
                marca_form = MarcaForm(request.POST, negocio=negocio)
                if marca_form.is_valid():
                    try:
                        marca = marca_form.save(commit=False)
                        marca.negocio = negocio
                        marca.save()
                        messages.success(request, "Marca registrada exitosamente.")
                    except IntegrityError as e:
                        if "unique_marca_negocio" in str(e):
                            error_message = "Ya existe una marca con este nombre en tu negocio."
                        else:
                            error_message = "Ocurrió un error al registrar la marca. Por favor, inténtalo nuevamente."
                else:
                    error_message = "Error en el formulario de marca. Revisa los datos ingresados."

            elif form_type == 'categoria':
                # Manejo del formulario de Categoría
                categoria_form = CategoriaForm(request.POST, negocio=negocio)
                if categoria_form.is_valid():
                    try:
                        categoria = categoria_form.save(commit=False)
                        categoria.negocio = negocio
                        categoria.save()
                        messages.success(request, "Categoría registrada exitosamente.")
                    except IntegrityError as e:
                        if "unique_categoria_negocio" in str(e):
                            error_message = "Ya existe una categoría con este nombre en tu negocio."
                        else:
                            error_message = "Ocurrió un error al registrar la categoría. Por favor, inténtalo nuevamente."
                else:
                    error_message = "Error en el formulario de categoría. Revisa los datos ingresados."

    return render(request, 'producto/list_prod.html', {
        'productos': productos,
        'estado_filtro': estado_filtro,
        'negocios': negocios,
        'negocio_filtro': negocio_filtro,
        'negocio_nombre': negocio_nombre,
        'categorias': categorias,
        'marcas': marcas,
        'categoria_filtro': categoria_filtro,
        'marca_form': marca_form,
        'categoria_form': categoria_form,
        'error_message': error_message,
        'palabra_clave': palabra_clave,
    })


@login_required
@permission_required('app.delete_producto', raise_exception=True)
def erase_prod(request, producto_id):
    producto = get_object_or_404(Producto, producto_id=producto_id)
    
    staff_profile = StaffProfile.objects.get(user=request.user)
    if producto.almacen.negocio != staff_profile.negocio:
        return HttpResponseForbidden("No tienes permiso para eliminar este producto.")

    if request.method == 'POST':
        # En lugar de eliminar el producto, cambiar su estado a 'sin_stock'
        producto.estado = 'sin_stock'
        producto.save()
        
        return redirect('list_prod')
    
    return redirect('list_prod')
#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
# Error de pago
def error_pago(request):
    return render(request, 'comprobante/pay_error.html')

def error_400(request, exception=None):
    return render(request, 'err/error_400.html', status=400)

def error_403(request, exception=None):
    return render(request, 'err/error_403.html', status=403)

def error_404(request, exception):
    return render(request, 'err/error_404.html', status=404)

def error_500(request):
    return render(request, 'err/error_500.html', status=500)

#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
#BODEGA
@login_required
@staff_member_required
@permission_required('app.view_entradabodega', raise_exception=True)
def historial_bodega(request):
    staff_profile = StaffProfile.objects.get(user=request.user)
    
    entradas = EntradaBodega.objects.filter(bodega__negocio=staff_profile.negocio).order_by('-fecha_recepcion')

    return render(request, 'bodega/historial_bodega.html', {'entradas': entradas})

@login_required
@permission_required('app.add_entradabodega', raise_exception=True)
def operaciones_bodega(request):
    staff_profile = StaffProfile.objects.get(user=request.user)
    almacen = Almacen.objects.filter(negocio=staff_profile.negocio).first()

    productos = Producto.objects.filter(almacen=almacen)
    proveedores = Proveedor.objects.filter(negocio=staff_profile.negocio)

    EntradaBodegaProductoFormSet = modelformset_factory(
        EntradaBodegaProducto,
        form=EntradaBodegaProductoForm,
        extra=1,
    )

    if request.method == 'POST':
        entrada_form = EntradaBodegaForm(request.POST, staff_profile=staff_profile)
        producto_formset = EntradaBodegaProductoFormSet(
            request.POST,
            queryset=EntradaBodegaProducto.objects.none(),
            form_kwargs={'almacen': almacen}
        )
        if entrada_form.is_valid() and producto_formset.is_valid():
            with transaction.atomic():
                entrada_bodega = entrada_form.save(commit=False)
                entrada_bodega.fecha_recepcion = request.POST.get("fecha_recepcion")
                entrada_bodega.bodega = almacen
                entrada_bodega.save()

                for producto_form in producto_formset:
                    if producto_form.cleaned_data:
                        entrada_producto = producto_form.save(commit=False)
                        entrada_producto.entrada_bodega = entrada_bodega

                        # Calcular precio_neto e iva_compra
                        entrada_producto.iva_compra = round(entrada_producto.precio_total * 0.19)
                        entrada_producto.precio_neto = entrada_producto.precio_total - entrada_producto.iva_compra

                        entrada_producto.save()

                        # Actualizar el stock del producto
                        producto = entrada_producto.producto
                        producto.stock += entrada_producto.cantidad_recibida
                        producto.save()

                messages.success(request, 'Entrada de bodega registrada exitosamente.')
            return redirect('operaciones_bodega')
    else:
        entrada_form = EntradaBodegaForm(staff_profile=staff_profile)
        producto_formset = EntradaBodegaProductoFormSet(
            queryset=EntradaBodegaProducto.objects.none(),
            form_kwargs={'almacen': almacen}
        )
    return render(request, 'bodega/operaciones_bodega.html', {
        'entrada_form': entrada_form,
        'producto_formset': producto_formset,
        'productos': productos,
        'proveedores': proveedores,
    })




@login_required
@permission_required('app.view_entradabodega', raise_exception=True)
def listar_entradas_bodega(request):
    # Obtener el perfil del staff y el negocio relacionado
    staff_profile = StaffProfile.objects.get(user=request.user)
    negocio = staff_profile.negocio

    # Filtrar entradas de bodega por el almacén del negocio
    entradas = EntradaBodega.objects.filter(bodega__negocio=negocio).order_by('-fecha_recepcion')

    # Renderizar el template con las entradas filtradas
    return render(request, 'bodega/lista_entradas_bodega.html', {
        'entradas': entradas,
    })



@login_required
@permission_required('app.view_entradabodega', raise_exception=True)
def detalle_entrada_bodega(request, entrada_id):
    # Obtener la entrada de bodega específica
    entrada = get_object_or_404(EntradaBodega, id=entrada_id)

    # Obtener los productos asociados a esta entrada, incluyendo las cantidades devueltas
    productos = EntradaBodegaProducto.objects.filter(entrada_bodega=entrada).annotate(
        subtotal=ExpressionWrapper(F('cantidad_recibida') * F('precio_total'), output_field=IntegerField()),
        cantidad_devuelta=Sum(
            'producto__devoluciones__cantidad_devuelta',
            filter=Q(producto__devoluciones__entrada_bodega=entrada),
            default=0
        ),
    ).annotate(
        cantidad_restante=F('cantidad_recibida') - F('cantidad_devuelta'),
        estado_devolucion=Case(
            When(cantidad_devuelta=0, then=Value('Sin Devolución')),
            When(cantidad_restante=0, then=Value('Devuelto Totalmente')),
            default=Value('Devuelto Parcialmente'),
            output_field=CharField(),
        )
    ).order_by('producto__nombre')

    # Contar el número total de productos
    total_productos = productos.count()
    # Calcular el total general
    total_general = productos.aggregate(total=Sum('subtotal'))['total'] or 0

    return render(request, 'bodega/detalle_entrada_bodega.html', {
        'entrada': entrada,
        'productos': productos,
        'total_productos': total_productos,
        'total_general': total_general,
    })



@login_required
@permission_required('app.view_entradabodega', raise_exception=True)
def detalle_prod_entrada_bodega(request, entrada_id, producto_id):
    # Obtener la entrada de bodega específica
    entrada = get_object_or_404(EntradaBodega, id=entrada_id)
    # Obtener el producto específico asociado a esta entrada
    producto_entrada = get_object_or_404(
        EntradaBodegaProducto,
        entrada_bodega=entrada,
        producto__producto_id=producto_id
    )
    
    return render(request, 'bodega/detalle_prod_entrada_bodega.html', {
        'entrada': entrada,
        'producto_entrada': producto_entrada,
    })


@login_required
@permission_required('app.add_productosdevueltos', raise_exception=True)
def devolver_factura(request, entrada_id):
    entrada = get_object_or_404(EntradaBodega, id=entrada_id)
    
    # Obtener los productos asociados a esta entrada
    productos_entrada = Producto.objects.filter(
        producto_id__in=EntradaBodegaProducto.objects.filter(entrada_bodega=entrada).values_list('producto_id', flat=True)
    )

    # Configuración del Formset con can_delete=True para permitir eliminar formularios
    DevolucionFormSet = modelformset_factory(
        ProductosDevueltos,
        form=DevolucionProductoForm,
        extra=1,
        #can_delete=True
    )

    if request.method == 'POST':
        formset = DevolucionFormSet(
            request.POST,
            queryset=ProductosDevueltos.objects.none(),
            form_kwargs={'productos_queryset': productos_entrada}
        )

        if formset.is_valid():
            with transaction.atomic():
                for form in formset:
                    if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                        devolucion = form.save(commit=False)
                        devolucion.entrada_bodega = entrada
                        devolucion.proveedor = entrada.proveedor

                        # Verificar cantidad disponible
                        producto_entrada = EntradaBodegaProducto.objects.get(
                            entrada_bodega=entrada,
                            producto=devolucion.producto
                        )
                        cantidad_devuelta_anterior = ProductosDevueltos.objects.filter(
                            entrada_bodega=entrada,
                            producto=devolucion.producto
                        ).aggregate(total=Sum('cantidad_devuelta'))['total'] or 0
                        cantidad_disponible = producto_entrada.cantidad_recibida - cantidad_devuelta_anterior

                        if devolucion.cantidad_devuelta > cantidad_disponible:
                            messages.error(
                                request,
                                f"La cantidad a devolver de {devolucion.producto.nombre} excede el disponible."
                            )
                            return redirect('devolver_factura', entrada_id=entrada.id)

                        devolucion.save()
                        devolucion.producto.stock -= devolucion.cantidad_devuelta
                        devolucion.producto.actualizar_estado()
                        devolucion.producto.save()

                messages.success(request, "Devolución registrada exitosamente.")
            return redirect('detalle_entrada_bodega', entrada_id=entrada.id)
    else:
        formset = DevolucionFormSet(
            queryset=ProductosDevueltos.objects.none(),
            form_kwargs={'productos_queryset': productos_entrada}
        )

    return render(request, 'bodega/devolver_factura.html', {
        'entrada': entrada,
        'formset': formset,
    })

@login_required
@permission_required('app.view_productosdevueltos', raise_exception=True)
def historial_devoluciones(request):
    devoluciones = ProductosDevueltos.objects.select_related('producto', 'entrada_bodega', 'proveedor').order_by('-fecha_devolucion')

    return render(request, 'bodega/historial_devoluciones.html', {
        'devoluciones': devoluciones,
    })


#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
#MANEJO DE STAFF
@login_required
@user_passes_test(lambda u: u.is_superuser)
def register_staff(request):
    grupos = Group.objects.all()  # Cargar los grupos disponibles una sola vez

    if request.method == 'POST':
        user_form = UserForm(request.POST)
        profile_form = StaffProfileForm(request.POST)

        if user_form.is_valid() and profile_form.is_valid():
            # Crear usuario
            user = user_form.save(commit=False)
            user.is_staff = True
            user.is_superuser = False
            user.is_active = True
            user.save()

            # Asignar grupo al usuario
            grupo_seleccionado = request.POST.get('grupo')
            if grupo_seleccionado:
                grupo = Group.objects.get(id=grupo_seleccionado)
                user.groups.add(grupo)

            # Crear perfil de staff
            staff_profile = profile_form.save(commit=False)
            staff_profile.user = user
            staff_profile.save()

            return redirect('list_staff')
    else:
        user_form = UserForm()
        profile_form = StaffProfileForm()

    return render(request, 'administration/staff/register_staff.html', {
        'user_form': user_form,
        'profile_form': profile_form,
        'grupos': grupos,  # Pasar todos los grupos disponibles
        'grupo_actual': None  # Ningún grupo está asignado aún
    })



@login_required
@user_passes_test(lambda u: u.is_superuser)
def list_staff(request):
    # Obtener el filtro de negocio desde la URL (si existe)
    negocio_filtro = request.GET.get('negocio', None)

    # Obtener todos los negocios para el dropdown
    negocios = Negocio.objects.all()

    # Si se ha seleccionado un negocio, filtrar los usuarios por ese negocio
    if negocio_filtro:
        try:
            negocio_filtro = int(negocio_filtro)
            staff_list = User.objects.filter(staffprofile__negocio_id=negocio_filtro, is_staff=True, is_active=True, is_superuser=False)
            negocio_nombre = Negocio.objects.get(id=negocio_filtro).nombre
        except (ValueError, Negocio.DoesNotExist):
            staff_list = User.objects.filter(is_staff=True, is_active=True, is_superuser=False)
            negocio_nombre = None
    else:
        # Si no hay filtro, mostrar todos los usuarios staff
        staff_list = User.objects.filter(is_staff=True, is_active=True, is_superuser=False)
        negocio_nombre = None

    return render(request, 'administration/staff/list_staff.html', {
        'staff_list': staff_list,
        'negocios': negocios,
        'negocio_filtro': negocio_filtro,
        'negocio_nombre': negocio_nombre
    })


@login_required
@user_passes_test(lambda u: u.is_superuser)
def mod_staff_account(request, staff_id):
    user = get_object_or_404(User, pk=staff_id)

    if request.method == 'POST':
        # Modificamos el nombre de usuario y contraseña
        user_form = UserForm(request.POST, instance=user)
        password_form = PasswordChangeForm(user, request.POST)

        if user_form.is_valid() and password_form.is_valid():
            user_form.save()

            # Cambiar la contraseña si se proporciona
            user = password_form.save()
            update_session_auth_hash(request, user)  # Mantener la sesión activa después del cambio de contraseña

            return redirect('list_staff')
    else:
        user_form = UserForm(instance=user)
        password_form = PasswordChangeForm(user)

    return render(request, 'administration/staff/mod_staff_account.html', {
        'user_form': user_form,
        'password_form': password_form,
        'staff': user,
    })


@login_required
@user_passes_test(lambda u: u.is_superuser)
def mod_staff_profile(request, staff_id):
    user = get_object_or_404(User, pk=staff_id)
    staff_profile = get_object_or_404(StaffProfile, user=user)

    grupos = Group.objects.all()

    if request.method == 'POST':
        profile_form = StaffProfileForm(request.POST, instance=staff_profile)

        if profile_form.is_valid():
            profile_form.save()

            # Asignar grupo al usuario
            grupo_seleccionado = request.POST.get('grupo')
            if grupo_seleccionado:
                grupo = Group.objects.get(id=grupo_seleccionado)
                user.groups.clear()
                user.groups.add(grupo)

            return redirect('list_staff')
    else:
        profile_form = StaffProfileForm(instance=staff_profile)

    return render(request, 'administration/staff/mod_staff_profile.html', {
        'profile_form': profile_form,
        'staff': user,
        'grupos': grupos,
        'grupo_actual': user.groups.first()
    })

#Borrar modificado para DESACTIVAR
@login_required
@user_passes_test(lambda u: u.is_superuser)
def erase_staff(request, staff_id):
    user = get_object_or_404(User, pk=staff_id)

    if request.method == 'POST':
        # Desactivar el perfil asociado
        StaffProfile.objects.filter(user=user).update(is_active=False)
        # Desactivar el usuario
        user.is_active = False
        user.save()
        return redirect('list_staff')

    return render(request, 'administration/staff/erase_staff.html', {'staff': user})
#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
#MANEJO DE ADMIN
# Listar administradores (solo superusuarios)
@login_required
@user_passes_test(lambda u: u.is_superuser)
def list_admin(request):
    admin_list = User.objects.filter(is_superuser=True) 
    return render(request, 'administration/admin/list_admin.html', {'admin_list': admin_list})

# Registrar nuevo administrador
@login_required
@user_passes_test(lambda u: u.is_superuser)
def register_admin(request):
    if request.method == 'POST':
        user_form = UserForm(request.POST)
        profile_form = StaffProfileForm(request.POST)
        
        if user_form.is_valid() and profile_form.is_valid():
            # Crear el usuario administrador
            user = user_form.save(commit=False)
            user.is_staff = True
            user.is_superuser = True
            user.is_active = True
            user.save()

            # Crear y vincular el perfil de Staff
            staff_profile = profile_form.save(commit=False)
            staff_profile.user = user
            staff_profile.save()

            return redirect('list_admin')
    else:
        user_form = UserForm()
        profile_form = StaffProfileForm()

    return render(request, 'administration/admin/register_admin.html', {
        'user_form': user_form,
        'profile_form': profile_form,
    })


@login_required
@user_passes_test(lambda u: u.is_superuser)
def mod_admin_profile(request, admin_id):
    admin = get_object_or_404(User, pk=admin_id, is_superuser=True)
    staff_profile = get_object_or_404(StaffProfile, user=admin)

    grupos = Group.objects.all()

    if request.method == 'POST':
        profile_form = StaffProfileForm(request.POST, instance=staff_profile)
        if profile_form.is_valid():
            profile_form.save()

            # Asignar grupo al usuario
            grupo_seleccionado = request.POST.get('grupo')
            if grupo_seleccionado:
                grupo = Group.objects.get(id=grupo_seleccionado)
                admin.groups.clear()
                admin.groups.add(grupo)
            return redirect('list_admin')
    else:
        profile_form = StaffProfileForm(instance=staff_profile)

    return render(request, 'administration/admin/mod_admin_profile.html', {
        'profile_form': profile_form,
        'admin': admin,
        'grupos': grupos,
        'grupo_actual': admin.groups.first()
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def mod_admin_account(request, admin_id):
    user = get_object_or_404(User, id=admin_id, is_superuser=True)
    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=user)
        if user_form.is_valid():
            user_form.save()
            return redirect('list_admin')
    else:
        user_form = UserForm(instance=user)
    return render(request, 'administration/admin/mod_admin_account.html', {'user_form': user_form})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def erase_admin(request, admin_id):
    user = get_object_or_404(User, pk=admin_id)

    if request.method == 'POST':
        StaffProfile.objects.filter(user=user).delete()
        user.delete()
        return redirect('list_admin')

    return render(request, 'administration/admin/erase_admin.html', {'admin': user})
#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
#Cambio de clave

# Cambiar la contraseña de un usuario
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Mantiene la sesión activa
            # Enviar correo de confirmación
            asunto = "¡Tu contraseña ha sido actualizada!"
            mensaje = (
                "Hola,\n\n"
                "Tu contraseña ha sido actualizada exitosamente. Si no realizaste este cambio, por favor "
                "contacta al jefe de tu negocio o informa de actividad sospechosa inmediatamente.\n\n"
                "Gracias por preferir Atlas Gestión."
            )
            enviar_correo_datos(
                destinatario=user.email,
                asunto=asunto,
                mensaje=mensaje
            )
            messages.success(request, 'Tu contraseña ha sido cambiada exitosamente.')
            return redirect('password_success')
        else:
            messages.error(request, 'Algo ocurrió, vuelve a intentarlo.')
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'administration/change_password.html', {'form': form})


@login_required
def password_success(request):
    return render(request, 'success/password_success.html')


@login_required
def licencia_vencida(request):
    return render(request, 'administration/licencia_vencida.html')
#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
#MANEJO DE NEGOCIOS
#Fx mixta - lista y creación negocio/almacén
@login_required
@user_passes_test(lambda u: u.is_superuser)
def list_negocios(request):
    negocios = Negocio.objects.all()  # Lista de negocios para mostrar

    if request.method == 'POST':
        form = NegocioForm(request.POST)
        if form.is_valid():
            # Crear el negocio pero no guardarlo aún
            negocio = form.save(commit=False)
            
            # Asignar las relaciones de región, provincia y comuna
            region_id = request.POST.get('region')
            provincia_id = request.POST.get('provincia')
            comuna_id = request.POST.get('comuna')
            
            negocio.region_id = region_id
            negocio.provincia_id = provincia_id
            negocio.comuna_id = comuna_id
            
            # Guardar el negocio
            negocio.save()
            
            almacen_direccion = form.cleaned_data['almacen_direccion']
            Almacen.objects.create(direccion=almacen_direccion, negocio=negocio)
            
            messages.success(request, 'Negocio y almacén creados exitosamente.')
            return redirect('list_negocios')
        else:
            for field, errors in form.errors.items():
                messages.error(request, f"Error en el campo {field}: {', '.join(errors)}")
            # Manejo de errores específicos
            if 'API' in str(errors):
                messages.error(request, "Hubo un problema al validar el RUT. Intenta nuevamente más tarde.")

    else:
        form = NegocioForm()

    regiones = Region.objects.all()

    return render(request, 'administration/negocio/list_negocios.html', {
        'form': form,
        'negocios': negocios,
        'regions': regiones,
    })


@login_required
@user_passes_test(lambda u: u.is_superuser)
def mod_negocio(request, negocio_id):
    negocio = get_object_or_404(Negocio, pk=negocio_id)
    almacen = Almacen.objects.filter(negocio=negocio).first() 

            
    if request.method == 'POST':
        form = ModNegocioForm(request.POST, request.FILES, instance=negocio)
        if form.is_valid():
            negocio = form.save(commit=False)
            negocio.save()
            
            if almacen:
                almacen.direccion = form.cleaned_data['almacen_direccion']
                almacen.save()
            return redirect('list_negocios')
    else:

        form = ModNegocioForm(instance=negocio, initial={
            'almacen_direccion': almacen.direccion if almacen else ''
        })

    return render(request, 'administration/negocio/mod_negocio.html', {
        'form': form,
    })


@login_required
@user_passes_test(lambda u: u.is_superuser)
def erase_negocio(request, negocio_id):
    negocio = get_object_or_404(Negocio, pk=negocio_id)
    
    if request.method == 'POST':
        negocio.delete()
        return redirect('list_negocios')

    return render(request, 'administration/negocio/erase_negocio.html', {'negocio': negocio})

# Método para cambiar el estado del negocio y las cuentas afiliadas
def cambiar_estado_negocio_y_cuentas(negocio_id, estado):
    try:
        negocio = Negocio.objects.get(id=negocio_id)
        negocio.is_active = estado  # Cambiar el estado del negocio
        negocio.save()

        # Cambiar el estado de las cuentas afiliadas
        usuarios_afiliados = User.objects.filter(staffprofile__negocio=negocio)
        usuarios_afiliados.update(is_active=estado)  # Sincronizar el estado con el negocio

        return True
    except Negocio.DoesNotExist:
        return False

@login_required
@user_passes_test(lambda u: u.is_superuser)
def cambiar_estado_negocio(request, negocio_id):
    try:
        negocio = Negocio.objects.get(id=negocio_id)
        nuevo_estado = not negocio.is_active  # Alternar el estado actual
        if cambiar_estado_negocio_y_cuentas(negocio_id, nuevo_estado):
            if nuevo_estado:
                messages.success(request, 'Negocio y cuentas afiliadas activados.')
            else:
                messages.success(request, 'Negocio y cuentas afiliadas inactivados.')
        else:
            messages.error(request, 'No se pudo cambiar el estado del negocio.')
    except Negocio.DoesNotExist:
        messages.error(request, 'El negocio no existe.')

    return redirect('list_negocios')


#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
#MANEJO DE PROVEEDORES
@login_required
@permission_required('app.add_proveedor', raise_exception=True)
def add_proveedor(request):
    try:
        staff_profile = StaffProfile.objects.get(user=request.user)
        negocio = staff_profile.negocio

        if request.method == 'POST':
            proveedor_form = ProveedorForm(request.POST)

            if proveedor_form.is_valid():
                proveedor = proveedor_form.save(commit=False)

                # Asignar las relaciones de región, provincia y comuna
                region_id = request.POST.get('region')
                provincia_id = request.POST.get('provincia')
                comuna_id = request.POST.get('comuna')

                proveedor.region = Region.objects.get(id=region_id) if region_id else None
                proveedor.provincia = Provincia.objects.get(id=provincia_id) if provincia_id else None
                proveedor.comuna = Comuna.objects.get(id=comuna_id) if comuna_id else None

                proveedor.negocio = negocio
                proveedor.save()

                messages.success(request, "Proveedor registrado exitosamente.")
                return redirect('list_proveedores')
        else:
            proveedor_form = ProveedorForm()
        
        regiones = Region.objects.all()

        return render(request, 'administration/proveedor/add_proveedor.html', {
            'proveedor_form': proveedor_form,
            'regions': regiones,
        })
    except StaffProfile.DoesNotExist:
        return HttpResponseForbidden("No tienes un perfil asociado.")


@login_required
@permission_required('app.view_proveedor', raise_exception=True)
def list_proveedores(request):
    try:
        staff_profile = StaffProfile.objects.get(user=request.user)
        negocio = staff_profile.negocio

        if request.user.is_superuser:
            negocios = Negocio.objects.all()
            proveedores = Proveedor.objects.all()
        else:
            negocios = None
            proveedores = Proveedor.objects.filter(negocio=negocio)

        negocio_filtro = request.GET.get('negocio', None)
        negocio_nombre = None
        if request.user.is_superuser and negocio_filtro:
            try:
                negocio_filtro = int(negocio_filtro)
                proveedores = proveedores.filter(negocio_id=negocio_filtro)
                negocio_nombre = Negocio.objects.get(id=negocio_filtro).nombre
            except (ValueError, Negocio.DoesNotExist):
                negocio_filtro = None

        if request.method == 'POST':
            form = ProveedorForm(request.POST)
            if form.is_valid():
                proveedor = form.save(commit=False)

                # Asignar las relaciones de región, provincia y comuna
                region_id = request.POST.get('region')
                provincia_id = request.POST.get('provincia')
                comuna_id = request.POST.get('comuna')

                proveedor.region = Region.objects.get(id=region_id) if region_id else None
                proveedor.provincia = Provincia.objects.get(id=provincia_id) if provincia_id else None
                proveedor.comuna = Comuna.objects.get(id=comuna_id) if comuna_id else None

                proveedor.negocio = negocio  
                proveedor.save()
                messages.success(request, "Proveedor registrado exitosamente.")
                return redirect('list_proveedores')
        else:
            form = ProveedorForm()

            regiones = Region.objects.all()

        return render(request, 'administration/proveedor/list_proveedores.html', {
            'proveedores': proveedores,
            'form': form,
            'negocios': negocios,
            'negocio_filtro': negocio_filtro,
            'negocio_nombre': negocio_nombre,
            'regions': regiones,
        })
    except StaffProfile.DoesNotExist:
        return HttpResponseForbidden("No tienes un perfil asociado.")

@login_required
@permission_required('app.change_proveedor', raise_exception=True)
def mod_proveedor(request, proveedor_id):
    try:
        proveedor = get_object_or_404(Proveedor, id=proveedor_id)

        if request.method == 'POST':
            proveedor_form = ProveedorForm(request.POST, instance=proveedor)

            if proveedor_form.is_valid():
                proveedor = proveedor_form.save(commit=False)

                region_id = request.POST.get('region')
                provincia_id = request.POST.get('provincia')
                comuna_id = request.POST.get('comuna')

                proveedor.region = Region.objects.get(id=region_id) if region_id else None
                proveedor.provincia = Provincia.objects.get(id=provincia_id) if provincia_id else None
                proveedor.comuna = Comuna.objects.get(id=comuna_id) if comuna_id else None

                proveedor.save()

                messages.success(request, "Proveedor modificado exitosamente.")
                return redirect('list_proveedores')
        else:
            proveedor_form = ProveedorForm(instance=proveedor)
            regiones = Region.objects.all()

        return render(request, 'administration/proveedor/mod_proveedor.html', {
            'proveedor_form': proveedor_form,
            'regions': regiones,
            'proveedor': proveedor,
        })
    except Proveedor.DoesNotExist:
        return HttpResponseForbidden("Proveedor no encontrado.")

@login_required
@permission_required('app.delete_proveedor', raise_exception=True)
def erase_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, pk=proveedor_id)
    staff_profile = StaffProfile.objects.get(user=request.user)

    if proveedor.negocio != staff_profile.negocio:
        return HttpResponseForbidden("No tienes permiso para eliminar este proveedor.")

    if request.method == 'POST':
        proveedor.delete()
        return redirect('list_proveedores')

    return render(request, 'administration/proveedor/erase_proveedor.html', {'proveedor': proveedor})

#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
#MANEJO DE CATEGORIAS
@login_required
@permission_required('app.view_categoria', raise_exception=True)
def list_categorias(request):
    try:
        # Determinar el negocio del usuario
        staff_profile = StaffProfile.objects.get(user=request.user)
        negocio = staff_profile.negocio

        if request.user.is_superuser:
            negocios = Negocio.objects.all()
            categorias = Categoria.objects.all()
        else:
            negocios = None
            categorias = Categoria.objects.filter(negocio=negocio)

        negocio_filtro = request.GET.get('negocio', None)
        negocio_nombre = None
        if request.user.is_superuser and negocio_filtro:
            try:
                negocio_filtro = int(negocio_filtro)
                categorias = categorias.filter(negocio_id=negocio_filtro)
                negocio_nombre = Negocio.objects.get(id=negocio_filtro).nombre
            except (ValueError, Negocio.DoesNotExist):
                negocio_filtro = None

        if request.method == 'POST':
            form = CategoriaForm(request.POST)
            if form.is_valid():
                categoria = form.save(commit=False)
                categoria.negocio = negocio
                categoria.save()
                messages.success(request, "Categoría registrada exitosamente.")
                return redirect('list_categorias')
        else:
            form = CategoriaForm()

        return render(request, 'categorias/list_categorias.html', {
            'categorias': categorias,
            'form': form,
            'negocios': negocios,
            'negocio_filtro': negocio_filtro,
            'negocio_nombre': negocio_nombre,
        })
    except StaffProfile.DoesNotExist:
        return HttpResponseForbidden("No tienes un perfil asociado.")


@login_required
@permission_required('app.add_categoria', raise_exception=True)
def add_categoria(request):
    try:
        staff_profile = StaffProfile.objects.get(user=request.user)
        negocio = staff_profile.negocio

        if request.method == 'POST':
            form = CategoriaForm(request.POST)
            if form.is_valid():
                categoria = form.save(commit=False)
                categoria.negocio = negocio  
                categoria.save()
                messages.success(request, "Categoría registrada exitosamente.")
                return redirect('list_categorias')
        else:
            form = CategoriaForm()

        return render(request, 'categorias/add_categoria.html', {
            'form': form,
        })
    except StaffProfile.DoesNotExist:
        return HttpResponseForbidden("No tienes un perfil asociado.")


@login_required
@permission_required('app.change_categoria', raise_exception=True)
def mod_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    staff_profile = StaffProfile.objects.get(user=request.user)

    if categoria.negocio != staff_profile.negocio:
        return HttpResponseForbidden("No tienes permiso para modificar esta categoría.")

    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            return redirect('list_prod')
    else:
        form = CategoriaForm(instance=categoria)

    return render(request, 'categorias/mod_categoria.html', {'form': form})

@login_required
@permission_required('app.delete_categoria', raise_exception=True)
def erase_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    staff_profile = StaffProfile.objects.get(user=request.user)

    if categoria.negocio != staff_profile.negocio:
        return HttpResponseForbidden("No tienes permiso para eliminar esta categoría.")

    if request.method == 'POST':
        categoria.delete()
        return redirect('list_prod')

    return render(request, 'categorias/erase_categoria.html', {'categoria': categoria})

#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
#MANEJO DE MARCAS
@login_required
@permission_required('app.add_marca', raise_exception=True)
def add_marca(request):
    try:
        # Determinar el negocio del usuario
        staff_profile = StaffProfile.objects.get(user=request.user)
        negocio = staff_profile.negocio

        error_message = None  # Variable para almacenar el mensaje de error

        if request.method == 'POST':
            form = MarcaForm(request.POST, negocio=negocio)
            if form.is_valid():
                try:
                    marca = form.save(commit=False)
                    marca.negocio = negocio  # Asociar la marca al negocio
                    marca.save()
                    messages.success(request, "Marca registrada exitosamente.")
                    return redirect('list_marcas')
                except IntegrityError as e:
                    # Captura el error de unicidad
                    if 'unique_marca_negocio' in str(e):  # Verifica el nombre de la restricción
                        error_message = "Ya existe una marca con este nombre en tu negocio."
                    else:
                        error_message = "Ocurrió un error al guardar la marca. Intenta nuevamente."
            else:
                error_message = "Error en el formulario. Verifica los datos ingresados."
        else:
            form = MarcaForm(negocio=negocio)

        # Renderiza la plantilla con el formulario y el mensaje de error (si existe)
        return render(request, 'marca/add_marca.html', {
            'form': form,
            'error_message': error_message,  # Pasamos el mensaje de error al contexto
        })
    except StaffProfile.DoesNotExist:
        return HttpResponseForbidden("No tienes un perfil asociado.")

@login_required
def list_marcas(request):
    try:
        staff_profile = StaffProfile.objects.get(user=request.user)
        negocio = staff_profile.negocio
    except StaffProfile.DoesNotExist:
        messages.error(request, "No tienes un negocio asignado.")
        return redirect('home')

    marcas = Marca.objects.filter(negocio=negocio)

    if request.method == 'POST':
        marca_form = MarcaForm(request.POST, negocio=negocio)
        if marca_form.is_valid():
            nueva_marca = marca_form.save(commit=False)
            nueva_marca.negocio = negocio
            nueva_marca.save()
            messages.success(request, "Marca registrada exitosamente.")
            return redirect('list_marcas')
        else:
            messages.error(request, "Error al registrar la marca. Verifica los datos.")
    else:
        marca_form = MarcaForm(negocio=negocio)

    return render(request, 'marca/list_marcas.html', {
        'marcas': marcas,
        'marca_form': marca_form,
    })


@login_required
def mod_marca(request, marca_id):
    try:
        staff_profile = StaffProfile.objects.get(user=request.user)
        negocio = staff_profile.negocio
    except StaffProfile.DoesNotExist:
        messages.error(request, "No tienes un negocio asignado.")
        return redirect('home')

    marca = get_object_or_404(Marca, id=marca_id, negocio=negocio)

    if request.method == 'POST':
        form = MarcaForm(request.POST, instance=marca, negocio=negocio)
        if form.is_valid():
            form.save()
            messages.success(request, "Marca modificada exitosamente.")
            return redirect('list_marcas')
        else:
            messages.error(request, "Error al modificar la marca. Verifica los datos.")
    else:
        form = MarcaForm(instance=marca, negocio=negocio)

    return render(request, 'marca/mod_marca.html', {'form': form, 'marca': marca})



@login_required
def erase_marca(request, marca_id):
    try:
        staff_profile = StaffProfile.objects.get(user=request.user)
        negocio = staff_profile.negocio
    except StaffProfile.DoesNotExist:
        messages.error(request, "No tienes un negocio asignado.")
        return redirect('home')

    marca = get_object_or_404(Marca, id=marca_id, negocio=negocio)

    if request.method == 'POST':
        marca.delete()
        messages.success(request, "Marca eliminada exitosamente.")
        return redirect('list_prod')

    return render(request, 'marca/erase_marca.html', {'marca': marca})


#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
# MANEJO DE TIPO DE PRODUCTO
@login_required
@permission_required('app.view_tipoproducto', raise_exception=True)
def list_tipos_producto(request):
    tipos = TipoProducto.objects.all()
    return render(request, 'tipoproducto/list_tipos_producto.html', {'tipos': tipos})

@login_required
@permission_required('app.add_tipoproducto', raise_exception=True)
def add_tipo_producto(request):
    if request.method == 'POST':
        form = TipoProductoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('list_tipos_producto')
    else:
        form = TipoProductoForm()
    return render(request, 'tipoproducto/add_tipo_producto.html', {'form': form})

@login_required
@permission_required('app.change_tipoproducto', raise_exception=True)
def mod_tipo_producto(request, tipo_id):
    tipo = get_object_or_404(TipoProducto, id=tipo_id)
    if request.method == 'POST':
        form = TipoProductoForm(request.POST, instance=tipo)
        if form.is_valid():
            form.save()
            return redirect('list_tipos_producto')
    else:
        form = TipoProductoForm(instance=tipo)
    return render(request, 'tipoproducto/mod_tipo_producto.html', {'form': form, 'tipo': tipo})

@login_required
@permission_required('app.delete_tipoproducto', raise_exception=True)
def erase_tipo_producto(request, tipo_id):
    tipo = get_object_or_404(TipoProducto, id=tipo_id)
    if request.method == 'POST':
        tipo.delete()
        return redirect('list_tipos_producto')
    return render(request, 'tipoproducto/erase_tipo_producto.html', {'tipo': tipo})


#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
# MANEJO DE PERFIL DE CLIENTES
@login_required
def add_cliente(request):
    if request.method == 'POST':
        correo = request.POST.get('correo')
        if correo:
            PerfilClientes.objects.get_or_create(correo=correo)
            return JsonResponse({'success': True, 'message': 'Cliente registrado'})
        return JsonResponse({'success': False, 'message': 'Correo inválido'})

@login_required
def list_clientes(request):
    clientes = PerfilClientes.objects.all()
    return render(request, 'administration/cuentas/list_cliente_modal.html', {'clientes': clientes})

@login_required
def mod_cliente(request, cliente_id):
    cliente = get_object_or_404(PerfilClientes, id=cliente_id)

    if request.method == 'POST':
        cliente_form = PerfilClientesForm(request.POST, instance=cliente)
        if cliente_form.is_valid():
            cliente_form.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({
                'success': False,
                'html_form': render_to_string('administration/cuentas/mod_cliente_modal.html', {'cliente_form': cliente_form, 'cliente': cliente}, request=request)
            })
    else:
        cliente_form = PerfilClientesForm(instance=cliente)

    html_form = render_to_string('administration/cuentas/mod_cliente_modal.html', {
        'cliente_form': cliente_form,
        'cliente': cliente,
    }, request=request)
    return JsonResponse({'html_form': html_form})

@login_required
def erase_cliente(request):
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente_id')
        cliente = get_object_or_404(PerfilClientes, id=cliente_id)
        cliente.delete()
        return JsonResponse({'success': True, 'message': 'Cliente eliminado'})

#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
# MANEJO DE PERFIL DE CLIENTES

""" Eventualmente descomentar *************
@login_required
def gestionar_empresas(request):
    staff_profile = StaffProfile.objects.get(user=request.user)
    negocio = staff_profile.negocio

    if request.method == 'POST':
        empresa_id = request.POST.get('empresa_id')
        accion = request.POST.get('accion')

        if accion == 'crear':
            form = PerfilClienteEmpresaForm(request.POST)
            if form.is_valid():
                empresa = form.save(commit=False)
                empresa.negocio = negocio
                empresa.activo = True
                                
                # Asignar las relaciones de región, provincia y comuna
                region_id = request.POST.get('region')
                provincia_id = request.POST.get('provincia')
                comuna_id = request.POST.get('comuna')

                empresa.region = Region.objects.get(id=region_id) if region_id else None
                empresa.provincia = Provincia.objects.get(id=provincia_id) if provincia_id else None
                empresa.comuna = Comuna.objects.get(id=comuna_id) if comuna_id else None

                empresa.save()
                return JsonResponse({'success': True, 'message': 'Empresa creada exitosamente.'})
            else:
                return JsonResponse({'success': False, 'message': 'Error al crear la empresa.'})

        elif accion == 'modificar':
            empresa = get_object_or_404(PerfilClienteEmpresa, id=empresa_id, negocio=negocio)
            form = PerfilClienteEmpresaForm(request.POST, instance=empresa)
            if form.is_valid():
                form.save()
                return JsonResponse({'success': True, 'message': 'Empresa modificada exitosamente.'})
            else:
                return JsonResponse({'success': False, 'message': 'Error al modificar la empresa.'})

        elif accion == 'eliminar':
            empresa = get_object_or_404(PerfilClienteEmpresa, id=empresa_id, negocio=negocio)
            empresa.activo = False
            empresa.save()
            return JsonResponse({'success': True, 'message': 'Empresa desactivada exitosamente.'})

    # Filtrar empresas activas
    empresas = PerfilClienteEmpresa.objects.filter(negocio=negocio, activo=True)
    form = PerfilClienteEmpresaForm()

    regiones = Region.objects.all()

    return render(request, 'caja/factura.html', {
        'empresas': empresas,
        'empresa_form': form,
        'regions': regiones
    })
 """



@login_required
@negocio_mayorista_required
def gestionar_empresas(request):
    staff_profile = StaffProfile.objects.get(user=request.user)
    empresas = PerfilClienteEmpresa.objects.filter(negocio=staff_profile.negocio, activo=True)
    empresa_form = PerfilClienteEmpresaForm()
    regiones = Region.objects.all()

    if request.method == 'POST':
        empresa_form = PerfilClienteEmpresaForm(request.POST)
        if empresa_form.is_valid():
            nueva_empresa = empresa_form.save(commit=False)
            nueva_empresa.negocio = staff_profile.negocio
            nueva_empresa.save()
            return redirect('gestionar_empresas')

    return render(request, 'caja/gestionar_empresas.html', {
        'empresas': empresas,
        'empresa_form': empresa_form,
        'regions': regiones
    })

@login_required
def editar_empresa(request, empresa_id):
    empresa = get_object_or_404(PerfilClienteEmpresa, id=empresa_id)
    
    if request.method == 'POST':
        form = PerfilClienteEmpresaForm(request.POST, instance=empresa)
        if form.is_valid():
            form.save()
            return redirect('gestionar_empresas')
    else:
        form = PerfilClienteEmpresaForm(instance=empresa)

    return render(request, 'caja/editar_empresa.html', {'form': form, 'empresa': empresa})

@login_required
def eliminar_empresa(request, empresa_id):
    empresa = get_object_or_404(PerfilClienteEmpresa, id=empresa_id)
    
    if request.method == 'POST':
        empresa.delete()
        messages.success(request, f"La empresa {empresa.nombre} ha sido eliminada con éxito.")
        return redirect('gestionar_empresas')

    messages.error(request, "Error al intentar eliminar la empresa.")
    return redirect('gestionar_empresas')




@login_required
def add_empresa(request):
    if request.method == 'POST':
        correo = request.POST.get('correo')
        if correo:
            PerfilClienteEmpresa.objects.get_or_create(correo=correo)
            return JsonResponse({'success': True, 'message': 'Cliente registrado'})
        return JsonResponse({'success': False, 'message': 'Correo inválido'})

@login_required
def list_empresa(request):
    clientes = PerfilClienteEmpresa.objects.all()
    return render(request, 'administration/empresas/list_empresa_modal.html', {'clientes': clientes})

@login_required
def mod_empresa(request, cliente_id):
    cliente = get_object_or_404(PerfilClienteEmpresa, id=cliente_id)

    if request.method == 'POST':
        cliente_form = PerfilClienteEmpresaForm(request.POST, instance=cliente)
        if cliente_form.is_valid():
            cliente_form.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({
                'success': False,
                'html_form': render_to_string('administration/empresas/mod_empresa_modal.html', {'cliente_form': cliente_form, 'cliente': cliente}, request=request)
            })
    else:
        cliente_form = PerfilClienteEmpresaForm(instance=cliente)

    html_form = render_to_string('administration/empresas/mod_empresa_modal.html', {
        'cliente_form': cliente_form,
        'cliente': cliente,
    }, request=request)
    return JsonResponse({'html_form': html_form})

@login_required
def erase_empresa(request):
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente_id')
        cliente = get_object_or_404(PerfilClienteEmpresa, id=cliente_id)
        cliente.delete()
        return JsonResponse({'success': True, 'message': 'Cliente eliminado'})


#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
# MANEJOS DE JEFE DE NEGOCIO - BOSS
# Registrar un nuevo staff (solo jefe de negocio)
@login_required
@jefe_required
def register_staff_for_boss(request):
    jefe_profile = StaffProfile.objects.get(user=request.user)
    negocio = jefe_profile.negocio

    if not negocio.puede_agregar_usuario():
        messages.error(request, "No puedes agregar más usuarios. Has alcanzado el límite de tu membresía.")
        return redirect('list_staff_for_boss')
    
    if request.method == 'POST':
        user_form = UserForBossForm(request.POST)
        profile_form = StaffProfileForBossForm(request.POST)

        try:
            if user_form.is_valid() and profile_form.is_valid():
                # Crear usuario con contraseña generada
                user = user_form.save(commit=False)
                user.is_staff = True

                # Generar contraseña y asignarla en los campos de confirmación
                passw = gen_password()
                user.set_password(passw)
                user_form.cleaned_data['password1'] = passw
                user_form.cleaned_data['password2'] = passw

                user.save()

                # Asignar grupo
                grupo_seleccionado = request.POST.get('grupo')
                if grupo_seleccionado:
                    grupo = Group.objects.get(id=grupo_seleccionado)
                    user.groups.add(grupo)

                # Crear perfil del staff
                staff_profile = profile_form.save(commit=False)
                staff_profile.user = user
                staff_profile.negocio = jefe_profile.negocio
                staff_profile.save()

                # Enviar credenciales al correo
                asunto = 'Credenciales de acceso'
                mensaje = (
                    f"Hola {user.first_name},\n\n"
                    f"Tu cuenta ha sido creada exitosamente.\n\n"
                    f"Credenciales de acceso:\n"
                    f"Usuario: {user.username}\n"
                    f"Contraseña: {passw}\n\n"
                    f"Recuerda cambiar tu contraseña en el primer inicio de sesión."
                )
                destinatarios = [user.email, request.user.email]

                for destinatario in destinatarios:
                    enviar_correo_datos(destinatario, asunto, mensaje)  

                return redirect('list_staff_for_boss')
            else:
                # Mostrar errores de validación
                raise ValueError("Los formularios no son válidos. Por favor, verifica los datos ingresados.")
        except Exception as e:
            # Capturar errores y enviarlos al template para su visualización
            error_message = str(e)
            return render(request, 'administration/for_boss/register_staff_for_boss.html', {
                'user_form': user_form,
                'profile_form': profile_form,
                'grupos': Group.objects.filter(name__in=['staff_bodega', 'staff_vendedor']),
                'error_message': error_message,
            })
    else:
        user_form = UserForBossForm()
        profile_form = StaffProfileForBossForm()

    return render(request, 'administration/for_boss/register_staff_for_boss.html', {
        'user_form': user_form,
        'profile_form': profile_form,
        'grupos': Group.objects.filter(name__in=['staff_bodega', 'staff_vendedor']),
    })




@login_required
@jefe_required
def mod_staff_profile_for_boss(request, staff_id):
    jefe_profile = StaffProfile.objects.get(user=request.user)

    # Validar que el usuario pertenece al negocio del jefe
    user = get_object_or_404(User, pk=staff_id)
    staff_profile = get_object_or_404(StaffProfile, user=user, negocio=jefe_profile.negocio)

    grupos = Group.objects.all()

    if request.method == 'POST':
        profile_form = StaffProfileForBossForm(request.POST, instance=staff_profile)

        if profile_form.is_valid():
            # Guardar cambios en el perfil
            profile_form.save()

            # Actualizar el grupo asignado al usuario
            grupo_seleccionado = request.POST.get('grupo')
            if grupo_seleccionado:
                grupo = get_object_or_404(Group, id=grupo_seleccionado)
                user.groups.clear()  # Limpiar todos los grupos actuales
                user.groups.add(grupo)  # Asignar el nuevo grupo

            return redirect('list_staff_for_boss')  # Redirigir a la lista de staff del jefe
    else:
        profile_form = StaffProfileForBossForm(instance=staff_profile)

    return render(request, 'administration/for_boss/mod_staff_profile_for_boss.html', {
        'profile_form': profile_form,
        'staff': user,
        'grupos': Group.objects.filter(name__in=['staff_bodega', 'staff_vendedor']),
        'grupo_actual': user.groups.first()  # Obtener el primer grupo asignado (si existe)
    })



# Listar staff (solo jefe de negocio)
from django.core.paginator import Paginator
@login_required
@jefe_required
def list_staff_for_boss(request):
    negocio = StaffProfile.objects.get(user=request.user).negocio
    staff_list = User.objects.filter(
        staffprofile__negocio=negocio,
        is_staff=True,
        is_active=True
    )
    paginator = Paginator(staff_list, 10)  # Mostrar 10 usuarios por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'administration/for_boss/list_staff_for_boss.html', {
        'staff_list_for_boss': page_obj
    })



# Modificar cuenta staff (solo jefe de negocio)
@jefe_required
def mod_staff_account_for_boss(request, staff_id):
    user = get_object_or_404(User, pk=staff_id)

    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=user)

        if user_form.is_valid():
            user_form.save()
            return redirect('list_staff_for_boss')
    else:
        user_form = UserForm(instance=user)

    return render(request, 'administration/for_boss/mod_staff_account_for_boss.html', {
        'user_form': user_form,
        'staff': user,
    })

# Eliminar staff (solo jefe de negocio) - DESACTIVAR 
@jefe_required
def erase_staff_for_boss(request, staff_id):
    user = get_object_or_404(User, pk=staff_id)

    if request.method == 'POST':
        # Desactivar el perfil asociado
        StaffProfile.objects.filter(user=user).update(is_active=False)
        # Desactivar el usuario
        user.is_active = False
        user.save()
        return redirect('list_staff_for_boss')

    return render(request, 'administration/for_boss/erase_staff_for_boss.html', {'staff': user})


#####################################
# / / / / / / / / / / / / / / / / / #
#####################################
# MANEJOS DE DASHBOARD 

def compras_por_mes(request):
    # Agrupar las compras por mes y por tipo de documento
    compras_boleta = (
        Compra.objects.filter(tipo_documento='boleta')
        .annotate(mes=TruncMonth('fecha'))
        .values('mes')
        .annotate(total=Sum('total'))
        .order_by('mes')
    )

    compras_factura = (
        Compra.objects.filter(tipo_documento='factura')
        .annotate(mes=TruncMonth('fecha'))
        .values('mes')
        .annotate(total=Sum('total'))
        .order_by('mes')
    )

    # Obtener los meses comunes
    meses = sorted({compra['mes'] for compra in compras_boleta}.union(
        {compra['mes'] for compra in compras_factura}
    ))

    # Formatear los datos
    labels = [mes.strftime('%B %Y') for mes in meses]
    data_boleta = [
        next((compra['total'] for compra in compras_boleta if compra['mes'] == mes), 0)
        for mes in meses
    ]
    data_factura = [
        next((compra['total'] for compra in compras_factura if compra['mes'] == mes), 0)
        for mes in meses
    ]

    return JsonResponse({
        'labels': labels,
        'data_boleta': data_boleta,
        'data_factura': data_factura,
    })


def compras_diarias(request):
    # Filtrar compras del mes actual
    ahora = datetime.now()
    compras_boleta = (
        Compra.objects.filter(fecha__year=ahora.year, fecha__month=ahora.month, tipo_documento='boleta')
        .annotate(dia=TruncDay('fecha'))
        .values('dia')
        .annotate(total=Sum('total'))
        .order_by('dia')
    )

    compras_factura = (
        Compra.objects.filter(fecha__year=ahora.year, fecha__month=ahora.month, tipo_documento='factura')
        .annotate(dia=TruncDay('fecha'))
        .values('dia')
        .annotate(total=Sum('total'))
        .order_by('dia')
    )

    # Obtener los días comunes
    dias = sorted({compra['dia'] for compra in compras_boleta}.union(
        {compra['dia'] for compra in compras_factura}
    ))

    # Formatear los datos
    labels = [dia.strftime('%d-%b-%Y') for dia in dias]
    data_boleta = [
        next((compra['total'] for compra in compras_boleta if compra['dia'] == dia), 0)
        for dia in dias
    ]
    data_factura = [
        next((compra['total'] for compra in compras_factura if compra['dia'] == dia), 0)
        for dia in dias
    ]

    return JsonResponse({
        'labels': labels,
        'data_boleta': data_boleta,
        'data_factura': data_factura,
    })

def generar_contexto_reporte(request):
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

    # Obtener el perfil del usuario actual
    user = request.user
    staff_profile = StaffProfile.objects.get(user=user)

    # Total ventas
    compras = Compra.objects.filter(usuario__staffprofile__negocio=staff_profile.negocio)
    total_ventas = compras.aggregate(Sum('total'))['total__sum'] or 0

    # Venta diaria
    today = now().date()
    venta_diaria = compras.filter(fecha__date=today).aggregate(
        total_hoy=Sum('total')
    )['total_hoy'] or 0

    # Margen de ventas
    total_entrada_bodega = EntradaBodegaProducto.objects.filter(
        entrada_bodega__proveedor__negocio=staff_profile.negocio
    ).aggregate(Sum('precio_total'))['precio_total__sum'] or 0

    margen_ventas = total_ventas - total_entrada_bodega

    # Porcentaje del margen de ventas
    porcentaje_margen = (margen_ventas / total_ventas * 100) if total_ventas > 0 else 0

    # Promedio ventas diario
    promedio_ventas_diario = compras.annotate(
        dia=TruncDay('fecha')
    ).values('dia').annotate(
        total_dia=Sum('total')
    ).aggregate(
        Avg('total_dia')
    )['total_dia__avg'] or 0

    # Producto más vendido
    producto_mas_vendido = DetalleCompra.objects.filter(
        compra__usuario__staffprofile__negocio=staff_profile.negocio
    ).values('producto__nombre').annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('-total_cantidad').first()

    # Producto menos vendido
    producto_menos_vendido = DetalleCompra.objects.filter(
        compra__usuario__staffprofile__negocio=staff_profile.negocio
    ).values('producto__nombre').annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('total_cantidad').first()

    # Número y porcentaje de compras realizadas con boleta y factura
    compras_por_tipo_documento = compras.values('tipo_documento').annotate(
        cantidad=Count('id')
    )
    total_compras = compras.count()
    for tipo in compras_por_tipo_documento:
        tipo['porcentaje'] = (tipo['cantidad'] / total_compras * 100) if total_compras > 0 else 0

    # Número y porcentaje de compras por método de pago
    compras_por_metodo_pago = compras.values('medio_pago').annotate(
        cantidad=Count('id')
    )
    for metodo in compras_por_metodo_pago:
        metodo['porcentaje'] = (metodo['cantidad'] / total_compras * 100) if total_compras > 0 else 0

    # Total de compras por mes diferenciando por tipo de documento
    compras_por_mes_boleta = (
        Compra.objects.filter(tipo_documento='boleta', usuario__staffprofile__negocio=staff_profile.negocio)
        .annotate(mes=TruncMonth('fecha'))
        .values('mes')
        .annotate(total=Sum('total'))
        .order_by('mes')
    )

    compras_por_mes_factura = (
        Compra.objects.filter(tipo_documento='factura', usuario__staffprofile__negocio=staff_profile.negocio)
        .annotate(mes=TruncMonth('fecha'))
        .values('mes')
        .annotate(total=Sum('total'))
        .order_by('mes')
    )

    # Total de compras diarias en el mes actual diferenciando por tipo de documento
    compras_diarias_boleta = (
        Compra.objects.filter(tipo_documento='boleta', usuario__staffprofile__negocio=staff_profile.negocio)
        .annotate(dia=TruncDay('fecha'))
        .values('dia')
        .annotate(total=Sum('total'))
        .order_by('dia')
    )

    compras_diarias_factura = (
        Compra.objects.filter(tipo_documento='factura', usuario__staffprofile__negocio=staff_profile.negocio)
        .annotate(dia=TruncDay('fecha'))
        .values('dia')
        .annotate(total=Sum('total'))
        .order_by('dia')
    )

    # Gráfico de barras: Compras por Mes
    meses = sorted(
        {compra['mes'] for compra in compras_por_mes_boleta}.union(
            {compra['mes'] for compra in compras_por_mes_factura}
        )
    )
    labels_mes = [mes.strftime('%B %Y') for mes in meses]
    totales_mes_boleta = [
        next((compra['total'] for compra in compras_por_mes_boleta if compra['mes'] == mes), 0)
        for mes in meses
    ]
    totales_mes_factura = [
        next((compra['total'] for compra in compras_por_mes_factura if compra['mes'] == mes), 0)
        for mes in meses
    ]

    fig1, ax1 = plt.subplots(figsize=(10, 6))
    width = 0.35
    x = range(len(labels_mes))
    ax1.bar(x, totales_mes_boleta, width, label="Boleta", color="skyblue", edgecolor="blue")
    ax1.bar(
        [i + width for i in x], totales_mes_factura, width, label="Factura", color="lightgreen", edgecolor="green"
    )
    ax1.set_title("Compras Totales por Mes (Boleta vs Factura)")
    ax1.set_ylabel("Total Compras (CLP)")
    ax1.set_xlabel("Mes")
    ax1.set_xticks([i + width / 2 for i in x])
    ax1.set_xticklabels(labels_mes, rotation=45, ha="right")
    ax1.legend()
    plt.tight_layout()

    buffer1 = BytesIO()
    plt.savefig(buffer1, format="png")
    buffer1.seek(0)
    grafico_barras = base64.b64encode(buffer1.getvalue()).decode("utf-8")
    buffer1.close()
    plt.close(fig1)

    # Gráfico de línea: Compras Diarias
    dias = sorted(
        {compra['dia'] for compra in compras_diarias_boleta}.union(
            {compra['dia'] for compra in compras_diarias_factura}
        )
    )
    labels_dia = [dia.strftime('%d-%b-%Y') for dia in dias]
    totales_dia_boleta = [
        next((compra['total'] for compra in compras_diarias_boleta if compra['dia'] == dia), 0)
        for dia in dias
    ]
    totales_dia_factura = [
        next((compra['total'] for compra in compras_diarias_factura if compra['dia'] == dia), 0)
        for dia in dias
    ]

    fig2, ax2 = plt.subplots(figsize=(12, 6))
    ax2.plot(labels_dia, totales_dia_boleta, marker='o', linestyle='-', color='blue', label="Boleta")
    ax2.plot(labels_dia, totales_dia_factura, marker='o', linestyle='-', color='green', label="Factura")
    ax2.fill_between(
        labels_dia, totales_dia_boleta, color="lightblue", alpha=0.5, label="Área Boleta"
    )
    ax2.fill_between(
        labels_dia, totales_dia_factura, color="lightgreen", alpha=0.5, label="Área Factura"
    )
    ax2.set_title("Compras Diarias del Mes Actual (Boleta vs Factura)")
    ax2.set_ylabel("Total Compras (CLP)")
    ax2.set_xlabel("Día")
    plt.xticks(rotation=45)
    plt.grid(True)
    plt.legend()
    plt.tight_layout()

    buffer2 = BytesIO()
    plt.savefig(buffer2, format="png")
    buffer2.seek(0)
    grafico_linea = base64.b64encode(buffer2.getvalue()).decode("utf-8")
    buffer2.close()
    plt.close(fig2)

    context = {
        # Métricas principales
        'total_compras': total_ventas,
        'promedio_diario_compras': promedio_ventas_diario,
        'producto_mas_comprado': producto_mas_vendido,
        'producto_menos_comprado': producto_menos_vendido,
        'compras_por_tipo_documento': compras_por_tipo_documento,
        'compras_por_metodo_pago': compras_por_metodo_pago,

        # Nuevas métricas
        'venta_diaria': venta_diaria,
        'margen_ventas': margen_ventas,
        'porcentaje_margen': porcentaje_margen,

        # Gráficos
        'grafico_barras': grafico_barras,
        'grafico_linea': grafico_linea,
    }

    return context


@login_required
def mostrar_reporte(request):
    context = generar_contexto_reporte(request) 
    return render(request, 'business/graficos.html', context)

@login_required
def generar_reporte_pdf(request):
    context = generar_contexto_reporte(request)

    # Renderizar el template HTML
    html_string = render_to_string('business/graficos.html', context)

    # Generar el PDF con WeasyPrint
    html = HTML(string=html_string)
    pdf = html.write_pdf()

    # Enviar como archivo adjunto
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_compras.pdf"'

    return response


from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from io import BytesIO
import base64
from PIL import Image as PILImage
import logging

logger = logging.getLogger(__name__)

def exportar_reportes_excel(request):
    try:
        # Obtener los datos del contexto usando la función `generar_contexto_reporte`
        contexto = generar_contexto_reporte(request)
        logger.debug(f"Contexto generado: {contexto}")

        # Crear un archivo Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Negocio"

        # Estilos
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Agregar Títulos
        sheet["A1"] = "Métricas del Negocio"
        sheet["A1"].font = bold_font
        sheet["A1"].alignment = center_alignment
        sheet.merge_cells("A1:B1")

        # Agregar Métricas Generales
        sheet.append(["Métrica", "Valor"])
        sheet.cell(row=2, column=1).font = bold_font
        sheet.cell(row=2, column=2).font = bold_font

        # Manejar 'producto_menos_comprado' de forma segura
        producto_menos = contexto.get('producto_menos_comprado')
        metrics = [
            ("Total de Compras", f"${contexto.get('total_compras', 0):,.0f}"),
            ("Promedio Diario de Compras", f"${contexto.get('promedio_diario_compras', 0):,.0f}"),
            (
                "Producto Más Vendido",
                f"{contexto['producto_mas_comprado']['producto__nombre']} ({contexto['producto_mas_comprado']['total_cantidad']} unidades)"
                if contexto.get('producto_mas_comprado') else "No disponible"
            ),
            (
                "Producto Menos Vendido",
                f"{producto_menos['producto__nombre']} ({producto_menos['total_cantidad']} unidades)"
                if producto_menos else "No disponible"
            )
        ]

        for metric in metrics:
            sheet.append(metric)

        # Agregar Compras por Tipo de Documento
        sheet.append([])  # Espacio
        sheet.append(["Compras por Tipo de Documento"])
        sheet.cell(sheet.max_row, 1).font = bold_font

        sheet.append(["Tipo de Documento", "Cantidad", "Porcentaje"])
        sheet.cell(sheet.max_row, 1).font = bold_font
        sheet.cell(sheet.max_row, 2).font = bold_font
        sheet.cell(sheet.max_row, 3).font = bold_font

        for tipo in contexto.get('compras_por_tipo_documento', []):
            sheet.append([
                tipo.get('tipo_documento', 'N/A'),
                tipo.get('cantidad', 0),
                f"{tipo.get('porcentaje', 0):.2f}%"
            ])

        # Agregar Compras por Método de Pago
        sheet.append([])  # Espacio
        sheet.append(["Compras por Método de Pago"])
        sheet.cell(sheet.max_row, 1).font = bold_font

        sheet.append(["Método de Pago", "Cantidad", "Porcentaje"])
        sheet.cell(sheet.max_row, 1).font = bold_font
        sheet.cell(sheet.max_row, 2).font = bold_font
        sheet.cell(sheet.max_row, 3).font = bold_font

        for metodo in contexto.get('compras_por_metodo_pago', []):
            sheet.append([
                metodo.get('medio_pago', 'N/A'),
                metodo.get('cantidad', 0),
                f"{metodo.get('porcentaje', 0):.2f}%"
            ])

        # Incluir gráficos en el Excel
        current_row = sheet.max_row + 2  # Dejar espacio después de las tablas

        # Gráfico de Barras
        grafico_barras = contexto.get('grafico_barras')
        if grafico_barras:
            grafico_barras_data = base64.b64decode(grafico_barras)
            grafico_barras_image = BytesIO(grafico_barras_data)
            barras_img = PILImage.open(grafico_barras_image)
            barras_img_path = "grafico_barras.png"
            barras_img.save(barras_img_path)  # Guardar temporalmente

            img_barras = Image(barras_img_path)
            img_barras.anchor = f"A{current_row}"  # Ubicar la imagen en la celda A
            sheet.add_image(img_barras)

        # Gráfico de Líneas
        current_row += 20  # Dejar espacio entre gráficos
        grafico_linea = contexto.get('grafico_linea')
        if grafico_linea:
            grafico_linea_data = base64.b64decode(grafico_linea)
            grafico_linea_image = BytesIO(grafico_linea_data)
            linea_img = PILImage.open(grafico_linea_image)
            linea_img_path = "grafico_linea.png"
            linea_img.save(linea_img_path)  # Guardar temporalmente

            img_linea = Image(linea_img_path)
            img_linea.anchor = f"A{current_row}"  # Ubicar la imagen en la celda A
            sheet.add_image(img_linea)

        # Ajustar ancho de columnas de forma segura
        for col_idx, col_cells in enumerate(sheet.columns, start=1):
            max_length = 0
            for cell in col_cells:
                if cell.value:  # Considerar solo celdas con contenido
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2  # Margen adicional
            column_letter = openpyxl.utils.get_column_letter(col_idx)  # Obtener la letra de la columna
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Exportar el archivo Excel
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="Reporte_Negocio.xlsx"'
        workbook.save(response)
        return response

    except Exception as e:
        logger.error(f"Error al exportar el reporte Excel: {e}", exc_info=True)
        return HttpResponse("Ocurrió un error al generar el reporte.", status=500)














import requests
from django.http import JsonResponse

def test_flask_connection(request):
    # URL de la API Flask
    flask_url = "http://127.0.0.1:5000/print"
    # Token de autorización
    token = "Bearer 1234abcd"
    
    # Datos a enviar
    payload = {
        "printer_name": "EPSON L3150 Series",
        "content": "Mensaje de prueba desde Django"
    }

    try:
        # Enviar la solicitud POST
        response = requests.post(
            flask_url,
            json=payload,
            headers={"Authorization": token}
        )
        
        # Devolver la respuesta de Flask como JSON
        return JsonResponse({
            "status_code": response.status_code,
            "response": response.json()
        })
    except requests.exceptions.RequestException as e:
        # Manejar errores de conexión
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)

from django.conf import settings 

def imprimir_boleta(printer_name, contenido):
    token = getattr(settings, 'ATLAS_PRINT_AGENT_TOKEN', 'Bearer 1234abcd')
    print_agent_url = "http://127.0.0.1:5000/print"

    payload = {
        "printer_name": printer_name,
        "content": contenido
    }
    headers = {
        "Authorization": token,
        "Content-Type": "application/json"
    }

    try:
        response = requests.post(print_agent_url, json=payload, headers=headers, timeout=5)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        return {"status": "error", "message": str(e)}


def generar_contenido_boleta(carrito, compra):
    detalles = carrito.carritoproducto_set.all()
    contenido = (
        f"Boleta #{compra.id}\n"
        f"Fecha: {compra.fecha}\n"
        f"Subtotal: {compra.subtotal:.2f}\n"
        f"Descuento: {compra.descuento_total:.2f}\n"
        f"IVA: {compra.iva_total:.2f}\n"
        f"Total: {compra.total:.2f}\n"
        "Detalles:\n"
    )
    for item in detalles:
        contenido += f"- {item.producto.nombre} x {item.cantidad} = {item.precio_unitario:.2f}\n"
    contenido += "\n¡Gracias por su compra!"
    return contenido
